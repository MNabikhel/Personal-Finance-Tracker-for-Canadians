"""Checks the built workbooks: their packages, their shape, and what the formulas say.

Two different things are being verified here.  The package tests confirm the
.xlsm really is a macro-enabled workbook - the content types, the relationship
to vbaProject.bin, the tables and defined names the macros address by name -
because a workbook that Excel refuses to open, or that opens with the macros
detached, fails no matter how good the VBA is.  The same tests confirm the
.xlsx edition is a plain workbook with none of that in it.

The recalculation tests then open each workbook in LibreOffice, throw away
every cached result and recompute from scratch, and compare what comes out
with the same totals worked out independently in Python from the sample
transactions.  LibreOffice is a separate implementation of both OOXML and the
spreadsheet function library, so agreement is real evidence the formulas are
right rather than a restatement of them.  For the edition without macros that
includes the categories themselves, which there are worked out by a formula
rather than written in by the importer.
"""

from __future__ import annotations

import io
import os
import sys
import tempfile
import unittest
import zipfile
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from typing import Dict, List

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import build as builder  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from tests import libreoffice  # noqa: E402
from tools import data, package, sample, workbook  # noqa: E402

TODAY = date(2026, 9, 1)
MONTH = workbook.report_month(TODAY)          # the month the workbook opens on

CENTS = 2

MACROS = "xlsm"          # the macro-enabled edition
PLAIN = "xlsx"           # the edition without macros

_blobs: Dict[str, bytes] = {}
_paths: Dict[str, str] = {}
_records: List[sample.Txn] = []


def setUpModule():
    global _records
    _records = sample.build(TODAY)
    _blobs[MACROS] = builder.build_package(TODAY)
    _blobs[PLAIN] = builder.build_plain_package(TODAY)
    for edition, blob in _blobs.items():
        handle, path = tempfile.mkstemp(prefix="cft-workbook-", suffix="." + edition)
        with os.fdopen(handle, "wb") as stream:
            stream.write(blob)
        _paths[edition] = path


def tearDownModule():
    for path in _paths.values():
        if os.path.exists(path):
            os.unlink(path)


# --- What the sample transactions add up to, worked out independently -------

TYPE_OF = {row[0]: row[2] for row in data.CATEGORIES}
GROUP_OF = {row[0]: row[1] for row in data.CATEGORIES}
ESSENTIAL_OF = {row[0]: row[3] for row in data.CATEGORIES}

# Which account belongs to whom, as the Accounts sheet lists it.  This is what
# the "Paid By" column works out from the account the money actually left.
PAID_BY = {
    sample.ACCOUNT_ALEX: sample.PERSON_A,
    sample.ACCOUNT_SAM: sample.PERSON_B,
    sample.ACCOUNT_JOINT: "Joint",
    sample.ACCOUNT_CARD: "Joint",
}

DEFAULT_SPLIT_A = Decimal("0.5")


def _round(value: Decimal) -> Decimal:
    """A spreadsheet ROUND: to the cent, with ties away from zero.

    Not Python's round(), which breaks ties towards the even digit and so
    disagrees with the sheet by a cent on exactly the half-cent amounts that
    splitting a bill in two produces.
    """
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _month(records: List[sample.Txn]) -> List[sample.Txn]:
    return [txn for txn in records if txn.month == MONTH]


def _of_type(records: List[sample.Txn], kind: str) -> List[sample.Txn]:
    return [txn for txn in records if TYPE_OF.get(txn.category, "Expense") == kind]


def _amount(txn: sample.Txn) -> Decimal:
    return Decimal(f"{txn.amount:.2f}")


def _split_a(txn: sample.Txn) -> Decimal:
    if txn.owner == sample.PERSON_A:
        return Decimal(1)
    if txn.owner == sample.PERSON_B:
        return Decimal(0)
    return DEFAULT_SPLIT_A


def _share_a(txn: sample.Txn) -> Decimal:
    return _round(_amount(txn) * _split_a(txn))


def _share_b(txn: sample.Txn) -> Decimal:
    return _round(_amount(txn) - _share_a(txn))


def _sum(values) -> float:
    return float(_round(sum(values, Decimal(0))))


def _total(records: List[sample.Txn]) -> float:
    return _sum(_amount(txn) for txn in records)


# --- Reading a recalculated copy of the workbook ----------------------------


class Recalculated:
    """The workbook open in LibreOffice, recomputed from the formulas alone."""

    def __init__(self, path: str):
        self.document = libreoffice.open_document(path)
        self.recalculate()

    def recalculate(self) -> None:
        self.document.calculateAll()

    def close(self) -> None:
        self.document.close(True)

    def cell(self, sheet: str, ref: str):
        return self.document.Sheets.getByName(sheet).getCellRangeByName(ref)

    def number(self, sheet: str, ref: str) -> float:
        return round(self.cell(sheet, ref).getValue(), CENTS)

    def text(self, sheet: str, ref: str) -> str:
        return self.cell(sheet, ref).getString()

    def set(self, sheet: str, ref: str, value: str) -> None:
        self.cell(sheet, ref).setString(value)
        self.recalculate()


class PackageTests(unittest.TestCase):
    """The parts of the package that make it a macro-enabled workbook."""

    def setUp(self):
        self.described = package.describe(_blobs[MACROS])

    def test_it_is_a_readable_zip_with_no_broken_members(self):
        for edition, blob in _blobs.items():
            with self.subTest(edition), zipfile.ZipFile(io.BytesIO(blob)) as archive:
                self.assertIsNone(archive.testzip())

    def test_the_vba_project_is_present_and_declared(self):
        self.assertIn("xl/vbaProject.bin", self.described["names"])
        self.assertTrue(self.described["vba_project"],
                        "the project part must not be empty")
        self.assertTrue(self.described["bin_default"],
                        "a .bin content type is needed or Excel drops the project")
        self.assertTrue(self.described["vba_relationship"],
                        "the workbook part must point at the project")

    def test_the_workbook_part_is_macro_enabled(self):
        # Without this the file is a .xlsx wearing a .xlsm extension and Excel
        # refuses to open it at all.
        self.assertTrue(self.described["macro_content_type"])
        self.assertFalse(self.described["sheet_content_type"],
                         "the plain spreadsheet content type must be replaced")

    def test_the_plain_edition_carries_no_trace_of_the_macros(self):
        # The other way round is just as fatal: a .xlsx that declares a VBA
        # project, or the macro-enabled content type, is a file Excel repairs
        # or refuses.
        described = package.describe(_blobs[PLAIN])
        self.assertNotIn("xl/vbaProject.bin", described["names"])
        self.assertFalse(described["vba_relationship"])
        self.assertFalse(described["macro_content_type"])
        self.assertTrue(described["sheet_content_type"])
        self.assertFalse(described["bin_default"])

    def test_two_builds_of_one_source_are_the_same_bytes(self):
        self.assertEqual(builder.build_package(TODAY), _blobs[MACROS])
        self.assertEqual(builder.build_plain_package(TODAY), _blobs[PLAIN])

    def test_the_document_is_dated_from_the_build_not_the_clock(self):
        for edition, blob in _blobs.items():
            with self.subTest(edition), zipfile.ZipFile(io.BytesIO(blob)) as archive:
                core = archive.read("docProps/core.xml").decode("utf-8")
                self.assertIn(f"{TODAY.isoformat()}T00:00:00Z", core)
                self.assertEqual(core.count(f"{TODAY.isoformat()}T00:00:00Z"), 2,
                                 "created and modified should agree")

    def test_the_app_name_and_version_match_the_ones_the_macros_use(self):
        # The name is the title of every message box the workbook shows and the
        # version is what the About message quotes, so either drifting apart
        # from the Settings sheet would be visible to the user.
        source = (os.path.join(os.path.dirname(os.path.dirname(
            os.path.abspath(__file__))), "vba", "modConst.bas"))
        with open(source, encoding="utf-8") as stream:
            text = stream.read()
        self.assertIn(f'APP_NAME As String = "{workbook.APP_NAME}"', text)
        self.assertIn(f'APP_VERSION As String = "{workbook.APP_VERSION}"', text)


class ShapeTests(unittest.TestCase):
    """The names the macros address the workbook by."""

    @classmethod
    def setUpClass(cls):
        cls.wb = load_workbook(io.BytesIO(_blobs[MACROS]), data_only=False)

    def test_every_sheet_the_macros_name_exists(self):
        self.assertEqual(set(workbook.CODE_NAMES) - set(self.wb.sheetnames), set())

    def test_sheet_code_names_survive_the_round_trip(self):
        # The VBA project's module stream binds a document module to a sheet by
        # code name; if these are lost the sheet modules bind to nothing.
        got = {name: self.wb[name].sheet_properties.codeName
               for name in workbook.CODE_NAMES}
        self.assertEqual(got, workbook.CODE_NAMES)

    def test_the_tables_the_macros_read_are_all_there(self):
        wanted = {
            workbook.SH_TXN: "tblTxn",
            workbook.SH_ACCOUNTS: "tblAccounts",
            workbook.SH_CATEGORIES: "tblCategories",
            workbook.SH_RULES: "tblRules",
            workbook.SH_FORMATS: "tblFormats",
            workbook.SH_LOG: "tblLog",
            workbook.SH_ENGINE: "tblTemplates",
        }
        for sheet, table in wanted.items():
            with self.subTest(table):
                self.assertIn(table, self.wb[sheet].tables)

    def test_the_ledger_has_every_column_the_macros_write(self):
        table = self.wb[workbook.SH_TXN].tables["tblTxn"]
        self.assertEqual([column.name for column in table.tableColumns],
                         workbook.TXN_HEADERS)

    def test_the_defined_names_the_macros_read_are_all_there(self):
        for key in ("ReportMonth", "ReportView", "PersonA", "PersonB",
                    "DefaultSplitA", "HouseholdMode", "Configured",
                    "TransferWindowDays", "SkipDuplicates", "CategoryList",
                    "AccountList", "FormatList", "TopMerchants", "TaxYear"):
            with self.subTest(key):
                self.assertIn(key, self.wb.defined_names)

    def test_the_formula_templates_match_the_built_rows(self):
        # The macros copy these onto rows they add, so a template that has
        # drifted from the built rows would quietly produce two kinds of row.
        engine = self.wb[workbook.SH_ENGINE]
        templates = {}
        for row in range(4, 4 + len(workbook.TXN_FORMULAS)):
            templates[engine[f"A{row}"].value] = engine[f"B{row}"].value
        self.assertEqual(templates, {header: workbook.formula_r1c1(header)
                                     for header in workbook.TXN_FORMULAS})

    def test_a_row_needing_a_category_is_flagged_across_its_whole_width(self):
        # A cellIs rule compares the cell being formatted, so spreading one
        # across the ledger only ever colours Category itself - which looks
        # right in the builder and does nothing on the sheet.
        ws = self.wb[workbook.SH_TXN]
        category = workbook.col_of("Category")
        first = workbook.col_of(workbook.TXN_HEADERS[0])
        last = workbook.col_of(workbook.TXN_HEADERS[-1])

        for ranges, rules in ws.conditional_formatting._cf_rules.items():
            if str(ranges.sqref).startswith(f"{first}") and last in str(ranges.sqref):
                for rule in rules:
                    if rule.type == "expression":
                        self.assertEqual(
                            rule.formula,
                            [f'${category}{workbook.TXN_FIRST_ROW + 1}'
                             f'="Uncategorized"'])
                        return
        self.fail("no whole-row rule for uncategorised transactions")

    def test_the_sample_data_is_actually_in_the_ledger(self):
        ws = self.wb[workbook.SH_TXN]
        first = workbook.TXN_FIRST_ROW + 1
        column = workbook.col_of("Amount")
        amounts = [ws[f"{column}{first + offset}"].value
                   for offset in range(len(_records))]
        self.assertEqual(amounts, [txn.amount for txn in _records])


class PlainShapeTests(unittest.TestCase):
    """What the edition without macros is made of."""

    @classmethod
    def setUpClass(cls):
        cls.wb = load_workbook(io.BytesIO(_blobs[PLAIN]), data_only=False)
        cls.ledger = cls.wb[workbook.SH_TXN]
        cls.first = workbook.TXN_FIRST_ROW + 1

    def test_the_importer_s_sheets_are_left_out(self):
        # Bank formats and an import log describe an importer this edition does
        # not have; every other sheet is there.
        expected = [name for name in workbook.CODE_NAMES
                    if name not in (workbook.SH_FORMATS, workbook.SH_LOG)]
        self.assertEqual(self.wb.sheetnames, expected)

    def test_the_ledger_has_the_same_columns_as_the_other_edition(self):
        # Rows must copy straight across between the two files.
        table = self.ledger.tables["tblTxn"]
        self.assertEqual([column.name for column in table.tableColumns],
                         workbook.TXN_HEADERS)

    def test_the_columns_the_importer_would_fill_are_formulas(self):
        for header in workbook.MANUAL_FORMULAS:
            with self.subTest(header):
                cell = self.ledger[f"{workbook.col_of(header)}{self.first}"]
                self.assertTrue(str(cell.value).startswith("="), cell.value)
                self.assertEqual(cell.value, workbook.formula_a1(
                    header, self.first, workbook.ledger_formulas(False)))

    def test_every_formula_column_is_declared_as_calculated(self):
        # This is what makes Excel fill the formulas into a row typed under the
        # table; without it the user gets a bare row and a broken ledger.
        table = self.ledger.tables["tblTxn"]
        declared = {column.name: column.calculatedColumnFormula.attr_text
                    for column in table.tableColumns
                    if column.calculatedColumnFormula is not None}
        formulas = workbook.ledger_formulas(False)
        self.assertEqual(set(declared), set(formulas))
        for header, text in declared.items():
            with self.subTest(header):
                self.assertEqual("=" + text,
                                 workbook.formula_a1(header, self.first, formulas))

    def test_ready_rows_wait_below_the_sample_data_inside_the_table(self):
        table = self.ledger.tables["tblTxn"]
        last = workbook.TXN_FIRST_ROW + len(_records) + workbook.READY_ROWS
        self.assertTrue(table.ref.endswith(f"{last}"), table.ref)
        row = workbook.TXN_FIRST_ROW + len(_records) + 1
        self.assertIsNone(self.ledger[f"{workbook.col_of('Date')}{row}"].value)
        self.assertIsNone(self.ledger[f"{workbook.col_of('Amount')}{row}"].value)
        for header in workbook.ledger_formulas(False):
            with self.subTest(header):
                cell = self.ledger[f"{workbook.col_of(header)}{row}"]
                self.assertTrue(str(cell.value).startswith("="), cell.value)

    def test_import_bookkeeping_is_hidden(self):
        for header in workbook.IMPORT_COLUMNS + workbook.TXN_HIDDEN:
            with self.subTest(header):
                self.assertTrue(
                    self.ledger.column_dimensions[workbook.col_of(header)].hidden)
        for header in ("Date", "Description", "Amount", "Category", "Owner"):
            with self.subTest(header):
                self.assertFalse(
                    self.ledger.column_dimensions[workbook.col_of(header)].hidden)

    def test_the_names_the_formulas_use_are_there_and_the_importer_s_are_not(self):
        for key in ("ReportMonth", "ReportView", "PersonA", "PersonB",
                    "DefaultSplitA", "HouseholdMode", "CategoryList",
                    "AccountList", "TaxYear"):
            with self.subTest(key):
                self.assertIn(key, self.wb.defined_names)
        self.assertNotIn("FormatList", self.wb.defined_names,
                         "a name pointing at a table that is not there")

    def test_the_help_and_the_notes_do_not_send_people_to_buttons(self):
        for sheet in (workbook.SH_HELP, workbook.SH_DASHBOARD, workbook.SH_SETTINGS,
                      workbook.SH_RULES, workbook.SH_REGISTERED, workbook.SH_TXN):
            ws = self.wb[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and not cell.value.startswith("="):
                        with self.subTest(sheet=sheet, cell=cell.coordinate):
                            self.assertNotIn("button", cell.value.lower())
                            self.assertNotIn("wizard", cell.value.lower())

    def test_new_functions_are_stored_with_the_prefix_excel_expects(self):
        # Excel stores functions newer than 2007 as _xlfn.NAME; written bare,
        # they show #NAME? until each cell is re-entered by hand.
        for header in workbook.ledger_formulas(False):
            cell = self.ledger[f"{workbook.col_of(header)}{self.first}"]
            self.assertNotRegex(cell.value, r"(?<![.\w])ISFORMULA\(", cell.value)


class RecalculationTests(unittest.TestCase):
    """Every number on the Dashboard, recomputed from the formulas."""

    sheet = workbook.SH_DASHBOARD
    edition = MACROS

    @classmethod
    def setUpClass(cls):
        try:
            libreoffice.context()
        except libreoffice.Unavailable as exc:  # pragma: no cover - environment
            raise unittest.SkipTest(str(exc))
        cls.book = Recalculated(_paths[cls.edition])

    @classmethod
    def tearDownClass(cls):
        cls.book.close()

    def test_it_opens_on_the_last_complete_month(self):
        self.assertEqual(self.book.text(self.sheet, "C6"), MONTH)
        self.assertEqual(self.book.text(self.sheet, "F6"), "Household")

    def test_money_in_out_and_saved(self):
        month = _month(_records)
        self.assertEqual(self.book.number(self.sheet, "C9"),
                         _total(_of_type(month, "Income")))
        self.assertEqual(self.book.number(self.sheet, "C10"),
                         -_total(_of_type(month, "Expense")))
        self.assertEqual(self.book.number(self.sheet, "C11"),
                         -_total(_of_type(month, "Saving")))

    def test_transfers_are_left_out_of_both_sides(self):
        # A credit card payment moves money without spending it.  Counting it
        # would double every card purchase, so it must appear in neither total.
        month = _month(_records)
        transfers = _of_type(month, "Transfer")
        self.assertTrue(transfers, "the sample month is meant to contain transfers")
        counted = (self.book.number(self.sheet, "C9")
                   + self.book.number(self.sheet, "C10")
                   + self.book.number(self.sheet, "C11"))
        everything = abs(_total(month))
        self.assertNotAlmostEqual(counted, everything, places=2)

    def test_net_cash_flow_and_savings_rate(self):
        money_in = self.book.number(self.sheet, "C9")
        money_out = self.book.number(self.sheet, "C10")
        saved = self.book.number(self.sheet, "C11")
        self.assertEqual(self.book.number(self.sheet, "C12"),
                         round(money_in - money_out - saved, CENTS))
        self.assertAlmostEqual(self.book.cell(self.sheet, "C13").getValue(),
                               (money_in - money_out) / money_in, places=6)

    def test_essential_and_discretionary_split_the_spending(self):
        month = _of_type(_month(_records), "Expense")
        essential = [txn for txn in month
                     if ESSENTIAL_OF.get(txn.category) == "Yes"]
        self.assertEqual(self.book.number(self.sheet, "F9"), -_total(essential))
        self.assertEqual(
            self.book.number(self.sheet, "F10"),
            round(self.book.number(self.sheet, "C10")
                  - self.book.number(self.sheet, "F9"), CENTS))

    def test_the_transaction_count(self):
        self.assertEqual(self.book.number(self.sheet, "F12"), len(_month(_records)))

    def test_nothing_is_left_uncategorised(self):
        self.assertEqual(self.book.number(self.sheet, "I6"), 0)

    def test_spending_by_group_accounts_for_everything_that_left(self):
        # The breakdown covers every group except Income and Transfers, so it
        # takes in savings as well as spending - money put into an RRSP did go
        # somewhere.  What it must not do is lose or double count a group.
        month = _month(_records)
        by_group: Dict[str, List[sample.Txn]] = {}
        for txn in month:
            if TYPE_OF.get(txn.category, "Expense") in ("Income", "Transfer"):
                continue
            by_group.setdefault(GROUP_OF.get(txn.category, "Other"), []).append(txn)

        got: Dict[str, float] = {}
        for offset, group in enumerate(data.SPENDING_GROUPS):
            got[group] = self.book.number(self.sheet, f"C{17 + offset}")
            with self.subTest(group):
                self.assertEqual(got[group], -_total(by_group.get(group, [])))

        self.assertEqual(
            round(sum(got.values()), CENTS),
            round(self.book.number(self.sheet, "C10")
                  + self.book.number(self.sheet, "C11"), CENTS),
            "the groups must account for the spending and the saving together")


class ReportsTests(unittest.TestCase):
    """The twelve months behind the Dashboard."""

    sheet = workbook.SH_REPORTS
    edition = MACROS

    @classmethod
    def setUpClass(cls):
        try:
            libreoffice.context()
        except libreoffice.Unavailable as exc:  # pragma: no cover - environment
            raise unittest.SkipTest(str(exc))
        cls.book = Recalculated(_paths[cls.edition])

    @classmethod
    def tearDownClass(cls):
        cls.book.close()

    def _column(self, index: int) -> str:
        from openpyxl.utils import get_column_letter
        return get_column_letter(workbook.REPORT_FIRST_COL + index)

    def test_the_twelve_months_end_on_the_report_month(self):
        months = [self.book.text(self.sheet, f"{self._column(index)}4")
                  for index in range(workbook.REPORT_MONTHS)]
        self.assertEqual(months[-1], MONTH)
        self.assertEqual(len(set(months)), workbook.REPORT_MONTHS)
        self.assertEqual(months, sorted(months), "months must read left to right")

    def test_each_month_column_agrees_with_the_transactions(self):
        for index in range(workbook.REPORT_MONTHS):
            column = self._column(index)
            month = self.book.text(self.sheet, f"{column}4")
            rows = [txn for txn in _records if txn.month == month]
            with self.subTest(month):
                self.assertEqual(self.book.number(self.sheet, f"{column}5"),
                                 _total(_of_type(rows, "Income")))
                self.assertEqual(self.book.number(self.sheet, f"{column}6"),
                                 -_total(_of_type(rows, "Expense")))
                self.assertEqual(self.book.number(self.sheet, f"{column}12"),
                                 len(rows))

    def test_the_total_column_sums_the_twelve(self):
        from openpyxl.utils import get_column_letter
        total = get_column_letter(workbook.REPORT_FIRST_COL + workbook.REPORT_MONTHS)
        for row in (5, 6, 7):
            with self.subTest(row=row):
                months = sum(self.book.number(self.sheet, f"{self._column(i)}{row}")
                             for i in range(workbook.REPORT_MONTHS))
                self.assertEqual(self.book.number(self.sheet, f"{total}{row}"),
                                 round(months, CENTS))


class CoupleModeTests(unittest.TestCase):
    """The part that only matters when two people share the workbook."""

    edition = MACROS

    @classmethod
    def setUpClass(cls):
        try:
            libreoffice.context()
        except libreoffice.Unavailable as exc:  # pragma: no cover - environment
            raise unittest.SkipTest(str(exc))
        cls.book = Recalculated(_paths[cls.edition])

    @classmethod
    def tearDownClass(cls):
        cls.book.close()

    def test_paid_by_follows_the_account_not_the_owner(self):
        # Someone can pay for something that is not theirs; that is the whole
        # reason there is anything to settle.
        month = _of_type(_month(_records), "Expense")
        for person, column in ((sample.PERSON_A, "C"), (sample.PERSON_B, "D")):
            paid = [txn for txn in month if PAID_BY[txn.account] == person]
            with self.subTest(person):
                self.assertEqual(self.book.number(workbook.SH_HOUSEHOLD, f"{column}7"),
                                 -_total(paid))

    def test_a_fair_share_divides_joint_spending_by_the_household_split(self):
        month = _of_type(_month(_records), "Expense")
        personal = [txn for txn in month
                    if PAID_BY[txn.account] in (sample.PERSON_A, sample.PERSON_B)]
        self.assertEqual(self.book.number(workbook.SH_HOUSEHOLD, "C8"),
                         -_sum(_share_a(txn) for txn in personal))
        self.assertEqual(self.book.number(workbook.SH_HOUSEHOLD, "D8"),
                         -_sum(_share_b(txn) for txn in personal))

    def test_the_settlement_says_who_owes_whom(self):
        balance = self.book.number(workbook.SH_HOUSEHOLD, "C9")
        settlement = self.book.text(workbook.SH_HOUSEHOLD, "C11")
        self.assertEqual(
            balance,
            round(self.book.number(workbook.SH_HOUSEHOLD, "C7")
                  - self.book.number(workbook.SH_HOUSEHOLD, "C8"), CENTS))
        if balance > 0:
            self.assertIn(f"{sample.PERSON_B} owes {sample.PERSON_A}", settlement)
        elif balance < 0:
            self.assertIn(f"{sample.PERSON_A} owes {sample.PERSON_B}", settlement)
        else:
            self.assertIn("Even", settlement)
        self.assertRegex(settlement, r"\$[\d,]+\.\d\d|Even")

    def test_the_two_balances_cancel(self):
        # Whatever one of them is owed, the other owes.  If these do not cancel
        # the shares are not a partition of the spending.
        self.assertEqual(
            round(self.book.number(workbook.SH_HOUSEHOLD, "C9")
                  + self.book.number(workbook.SH_HOUSEHOLD, "D9"), CENTS), 0.0)

    def test_the_view_selector_switches_every_report_to_one_person(self):
        month = _month(_records)
        try:
            for person, share in ((sample.PERSON_A, _share_a),
                                  (sample.PERSON_B, _share_b)):
                self.book.set(workbook.SH_DASHBOARD, "F6", person)
                with self.subTest(person):
                    self.assertEqual(
                        self.book.number(workbook.SH_DASHBOARD, "C9"),
                        _sum(share(txn) for txn in _of_type(month, "Income")))
                    self.assertEqual(
                        self.book.number(workbook.SH_DASHBOARD, "C10"),
                        -_sum(share(txn) for txn in _of_type(month, "Expense")))
        finally:
            self.book.set(workbook.SH_DASHBOARD, "F6", "Household")

    def test_the_two_personal_views_add_up_to_the_household_one(self):
        # Share A and Share B are meant to be a partition of the amount, so
        # neither person's view can lose or invent money.
        totals = {}
        try:
            for view in ("Household", sample.PERSON_A, sample.PERSON_B):
                self.book.set(workbook.SH_DASHBOARD, "F6", view)
                totals[view] = (self.book.number(workbook.SH_DASHBOARD, "C9"),
                                self.book.number(workbook.SH_DASHBOARD, "C10"))
        finally:
            self.book.set(workbook.SH_DASHBOARD, "F6", "Household")

        for index, label in enumerate(("money in", "money out")):
            with self.subTest(label):
                self.assertAlmostEqual(
                    totals[sample.PERSON_A][index] + totals[sample.PERSON_B][index],
                    totals["Household"][index], places=1)


# --- The edition without macros ---------------------------------------------
#
# Every report is expected to say exactly what it says in the other edition.
# That is a stronger claim here than it looks: the categories those reports
# add up are not written into the file, they are worked out by the Category
# formula from the Rules sheet, so agreement means the formula categorises the
# whole sample the way the importer does.

class PlainRecalculationTests(RecalculationTests):
    edition = PLAIN


class PlainReportsTests(ReportsTests):
    edition = PLAIN


class PlainCoupleModeTests(CoupleModeTests):
    edition = PLAIN


class FormulaCategoriserTests(unittest.TestCase):
    """The ledger columns that are formulas only in the edition without macros."""

    @classmethod
    def setUpClass(cls):
        try:
            libreoffice.context()
        except libreoffice.Unavailable as exc:  # pragma: no cover - environment
            raise unittest.SkipTest(str(exc))
        cls.book = Recalculated(_paths[PLAIN])
        cls.first = workbook.TXN_FIRST_ROW + 1
        cls.ready = cls.first + len(_records)      # the first row waiting empty

    @classmethod
    def tearDownClass(cls):
        cls.book.close()

    def _cell(self, header: str, row: int) -> str:
        return f"{workbook.col_of(header)}{row}"

    def _ledger(self, header: str, row: int) -> str:
        return self.book.text(workbook.SH_TXN, self._cell(header, row))

    def _type(self, row: int, **values) -> None:
        for header, value in values.items():
            cell = self.book.cell(workbook.SH_TXN, self._cell(header, row))
            if isinstance(value, str):
                cell.setString(value)
            elif isinstance(value, date):
                # A spreadsheet date is days since 1899-12-30.
                cell.setValue((value - date(1899, 12, 30)).days)
            else:
                cell.setValue(value)
        self.book.recalculate()

    def _clear(self, row: int) -> None:
        for header in ("Date", "Account", "Description", "Amount"):
            self.book.cell(workbook.SH_TXN, self._cell(header, row)).setString("")
        self.book.recalculate()

    def test_every_sample_row_gets_the_category_the_importer_gives_it(self):
        got = [self._ledger("Category", self.first + offset)
               for offset in range(len(_records))]
        wanted = [txn.category for txn in _records]
        wrong = [(txn.description, want, have)
                 for txn, want, have in zip(_records, wanted, got) if want != have]
        self.assertEqual(wrong, [], f"{len(wrong)} of {len(_records)} differ")

    def test_every_sample_row_gets_the_owner_the_importer_gives_it(self):
        got = [self._ledger("Owner", self.first + offset)
               for offset in range(len(_records))]
        self.assertEqual(got, [txn.owner for txn in _records])

    def test_rows_are_numbered_and_tagged_as_the_rules_work(self):
        self.assertEqual(self._ledger("Txn ID", self.first), "T000001")
        self.assertEqual(self._ledger("Txn ID", self.first + len(_records) - 1),
                         f"T{len(_records):06d}")
        self.assertEqual(self._ledger("Tagged By", self.first), "Rule")

    def test_a_row_waiting_to_be_filled_shows_nothing(self):
        for header in workbook.ledger_formulas(False):
            with self.subTest(header):
                self.assertEqual(self._ledger(header, self.ready), "")
        self.assertEqual(self.book.number(workbook.SH_DASHBOARD, "I6"), 0,
                         "empty rows must not count as needing a category")

    def test_typing_a_transaction_fills_the_rest_of_the_row_in(self):
        row = self.ready
        try:
            self._type(row, Date=date(2026, 8, 3), Account=sample.ACCOUNT_ALEX,
                       Description="IDP PURCHASE - 4411 LOBLAWS #1077 TORONTO ON",
                       Amount=-62.30)
            self.assertEqual(self._ledger("Txn ID", row), f"T{len(_records) + 1:06d}")
            self.assertEqual(self._ledger("Month", row), "2026-08")
            self.assertEqual(self._ledger("Category", row), "Groceries")
            self.assertEqual(self._ledger("Group", row), "Food")
            self.assertEqual(self._ledger("Paid By", row), sample.PERSON_A)
            # Groceries are the household's whoever's card paid for them.
            self.assertEqual(self._ledger("Owner", row), "Joint")
            self.assertEqual(self._ledger("Tagged By", row), "Rule")
            self.assertEqual(self.book.number(workbook.SH_TXN, self._cell("Share A", row)),
                             -31.15)
        finally:
            self._clear(row)

    def test_a_description_no_rule_knows_is_flagged_not_guessed(self):
        row = self.ready
        try:
            self._type(row, Date=date(2026, 8, 4), Account=sample.ACCOUNT_SAM,
                       Description="ZORBLAX HOLDINGS 9921", Amount=-19.99)
            self.assertEqual(self._ledger("Category", row), "Uncategorized")
            self.assertEqual(self.book.number(workbook.SH_DASHBOARD, "I6"), 1)
        finally:
            self._clear(row)

    def test_flow_is_honoured_so_a_refund_is_not_a_purchase(self):
        # Every merchant rule is "Money out"; the same text with the money
        # coming back is not a grocery bill.
        row = self.ready
        try:
            self._type(row, Date=date(2026, 8, 5), Account=sample.ACCOUNT_CARD,
                       Description="LOBLAWS #1077 TORONTO ON", Amount=48.20)
            self.assertEqual(self._ledger("Category", row), "Uncategorized")
        finally:
            self._clear(row)

    def test_a_word_rule_does_not_fire_inside_a_longer_word(self):
        row = self.ready
        try:
            self._type(row, Date=date(2026, 8, 6), Account=sample.ACCOUNT_ALEX,
                       Description="MOBIL 4402 GAS BAR ETOBICOKE ON", Amount=-71.00)
            self.assertEqual(self._ledger("Category", row), "Fuel")
            self._type(row, Description="FREEDOM MOBILE PREAUTHORIZED DEBIT")
            self.assertEqual(self._ledger("Category", row), "Mobile Phone")
        finally:
            self._clear(row)

    def test_priority_decides_between_two_rules_that_both_match(self):
        # "TRANSFER TO" says Internal Transfer at priority 12; "RRSP" says
        # RRSP Contribution at priority 9 and must win, or the year's room is
        # never counted.  The rows are in priority order on the sheet, so this
        # is also the check that the formula takes the first match.
        row = self.ready
        try:
            self._type(row, Date=date(2026, 8, 7), Account=sample.ACCOUNT_ALEX,
                       Description="TRANSFER TO RRSP WEALTHSIMPLE", Amount=-450)
            self.assertEqual(self._ledger("Category", row), "RRSP Contribution")
        finally:
            self._clear(row)

    def test_typing_over_the_category_marks_the_row_manual(self):
        row = self.first
        formula = self.book.cell(workbook.SH_TXN, self._cell("Category", row)).getFormula()
        try:
            self.book.set(workbook.SH_TXN, self._cell("Category", row), "Hobbies")
            self.assertEqual(self._ledger("Tagged By", row), "Manual")
            self.assertEqual(self._ledger("Group", row), "Personal")
        finally:
            self.book.cell(workbook.SH_TXN, self._cell("Category", row)).setFormula(formula)
            self.book.recalculate()
        self.assertEqual(self._ledger("Tagged By", row), "Rule")

    def test_disabling_a_rule_on_the_rules_sheet_takes_effect_at_once(self):
        rules = self.book.document.Sheets.getByName(workbook.SH_RULES)
        patterns = [rules.getCellByPosition(5, 4 + index).getString()
                    for index in range(len(data.seed_rules()))]
        index = patterns.index("TIM HORTONS")
        enabled = rules.getCellByPosition(2, 4 + index)
        coffee = next(offset for offset, txn in enumerate(_records)
                      if txn.category == "Coffee & Snacks"
                      and "TIM HORTONS" in txn.description)
        row = self.first + coffee
        try:
            enabled.setString("No")
            self.book.recalculate()
            self.assertNotEqual(self._ledger("Category", row), "Coffee & Snacks")
        finally:
            enabled.setString("Yes")
            self.book.recalculate()
        self.assertEqual(self._ledger("Category", row), "Coffee & Snacks")

    def test_the_dashboard_ranks_the_month_s_categories(self):
        month = _month(_records)
        by_category: Dict[str, Decimal] = {}
        for txn in month:
            if TYPE_OF.get(txn.category, "Expense") in ("Income", "Transfer"):
                continue
            by_category[txn.category] = by_category.get(txn.category, Decimal(0)) \
                - _amount(txn)
        ranked = sorted(by_category.items(), key=lambda item: -item[1])

        top = workbook.DASH_MERCHANTS_TOP
        for place, (category, spent) in enumerate(ranked[:10], start=1):
            with self.subTest(place=place):
                self.assertEqual(
                    self.book.text(workbook.SH_DASHBOARD, f"B{top + 1 + place}"),
                    category)
                self.assertEqual(
                    self.book.number(workbook.SH_DASHBOARD, f"C{top + 1 + place}"),
                    float(_round(spent)))
        shares = sum(self.book.cell(workbook.SH_DASHBOARD, f"E{top + 1 + place}").getValue()
                     for place in range(1, 11))
        self.assertLessEqual(shares, 1.0 + 1e-9)


if __name__ == "__main__":
    unittest.main()
