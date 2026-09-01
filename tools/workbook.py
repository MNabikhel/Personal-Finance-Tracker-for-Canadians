"""Builds the workbook body (sheets, tables, formulas, charts, validation).

The macros never invent numbers: every figure on the Dashboard, Reports, Budget,
Tax Summary and Household sheets is a live worksheet formula, so the workbook is
still readable with macros switched off.
"""

from __future__ import annotations

import re
from datetime import date, datetime, time, timedelta
from typing import Iterable, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.formatting.rule import (CellIsRule, ColorScaleRule, DataBarRule,
                                      FormulaRule)
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.defined_name import DefinedName

from . import data, sample

APP_NAME = "Canadian Finance Tracker"       # matches modConst.APP_NAME

# --- Look and feel ----------------------------------------------------------

RED = "C41E3A"
RED_DARK = "8E1428"
INK = "1F2933"
MUTED = "6B7280"
PANEL = "F4F5F7"
ACCENT = "0F6E5C"
WHITE = "FFFFFF"

MONEY = '#,##0.00;[Red]-#,##0.00'
MONEY0 = '#,##0;[Red]-#,##0'
PERCENT = '0.0%'
DATE_FMT = 'yyyy-mm-dd'

TITLE_FONT = Font(name="Calibri", size=20, bold=True, color=RED)
SUB_FONT = Font(name="Calibri", size=10, color=MUTED, italic=True)
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color=WHITE)
LABEL_FONT = Font(name="Calibri", size=11, color=INK)
BOLD = Font(name="Calibri", size=11, bold=True, color=INK)
KPI_FONT = Font(name="Calibri", size=14, bold=True, color=INK)
NOTE_FONT = Font(name="Calibri", size=9, color=MUTED)
HEAD_FILL = PatternFill("solid", fgColor=RED)
PANEL_FILL = PatternFill("solid", fgColor=PANEL)
THIN = Side(style="thin", color="D9D9E3")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# --- Sheet names ------------------------------------------------------------

SH_DASHBOARD = "Dashboard"
SH_TXN = "Transactions"
SH_ACCOUNTS = "Accounts"
SH_CATEGORIES = "Categories"
SH_RULES = "Rules"
SH_BUDGET = "Budget"
SH_REPORTS = "Reports"
SH_HOUSEHOLD = "Household"
SH_TAX = "Tax Summary"
SH_REGISTERED = "Registered Plans"
SH_FORMATS = "Bank Formats"
SH_SETTINGS = "Settings"
SH_LOG = "Import Log"
SH_HELP = "Help"
SH_ENGINE = "Engine"

CODE_NAMES = {
    SH_DASHBOARD: "shDashboard",
    SH_TXN: "shTransactions",
    SH_ACCOUNTS: "shAccounts",
    SH_CATEGORIES: "shCategories",
    SH_RULES: "shRules",
    SH_BUDGET: "shBudget",
    SH_REPORTS: "shReports",
    SH_HOUSEHOLD: "shHousehold",
    SH_TAX: "shTax",
    SH_REGISTERED: "shRegistered",
    SH_FORMATS: "shFormats",
    SH_SETTINGS: "shSettings",
    SH_LOG: "shLog",
    SH_HELP: "shHelp",
    SH_ENGINE: "shEngine",
}

TAB_COLOURS = {
    SH_DASHBOARD: RED,
    SH_TXN: "2F5597",
    SH_REPORTS: ACCENT,
    SH_BUDGET: ACCENT,
    SH_HOUSEHOLD: ACCENT,
    SH_TAX: "7F5A18",
    SH_REGISTERED: "7F5A18",
}

# --- Transactions table -----------------------------------------------------

TXN_HEADERS = [
    "Txn ID",
    "Date",
    "Month",
    "Account",
    "Paid By",
    "Owner",
    "Description",
    "Merchant",
    "Amount",
    "Category",
    "Group",
    "Type",
    "Essential",
    "Tax Tag",
    "Split A %",
    "Share A",
    "Share B",
    "Reimbursable",
    "Notes",
    "Source File",
    "Batch",
    "Match Key",
    "Tagged By",
    "View Amount",
]

TXN_WIDTHS = {
    "Txn ID": 10, "Date": 11, "Month": 9, "Account": 22, "Paid By": 12,
    "Owner": 12, "Description": 46, "Merchant": 24, "Amount": 12,
    "Category": 24, "Group": 14, "Type": 10, "Essential": 10, "Tax Tag": 18,
    "Split A %": 9, "Share A": 12, "Share B": 12, "Reimbursable": 12,
    "Notes": 24, "Source File": 22, "Batch": 18, "Match Key": 12,
    "Tagged By": 12, "View Amount": 12,
}

# Plumbing rather than data: hidden so the ledger stays readable.
TXN_HIDDEN = ["View Amount"]

# Calculated columns.  @{Header} stands for "this row, that column"; it is
# expanded to a plain relative reference for the built rows and to R1C1 for the
# templates the macros use when they add rows.
#
# Plain references are used instead of [@Header] structured references on
# purpose: they mean exactly the same thing to Excel, but they also evaluate in
# LibreOffice Calc, which does not implement the [@...] form.
TXN_FORMULAS = {
    "Month": 'IF(@{Date}="","",TEXT(@{Date},"yyyy-mm"))',
    "Paid By": ('IFERROR(INDEX(tblAccounts[Owner],'
                'MATCH(@{Account},tblAccounts[Account],0)),"Joint")'),
    "Group": ('IFERROR(INDEX(tblCategories[Group],'
              'MATCH(@{Category},tblCategories[Category],0)),"Other")'),
    "Type": ('IFERROR(INDEX(tblCategories[Type],'
             'MATCH(@{Category},tblCategories[Category],0)),'
             'IF(@{Amount}>0,"Income","Expense"))'),
    "Essential": ('IFERROR(INDEX(tblCategories[Essential],'
                  'MATCH(@{Category},tblCategories[Category],0)),"No")'),
    "Tax Tag": ('IFERROR(INDEX(tblCategories[Tax Tag],'
                'MATCH(@{Category},tblCategories[Category],0)),"")'),
    "Split A %": (
        'IF(@{Date}="","",IF(@{Owner}=PersonA,1,IF(@{Owner}=PersonB,0,'
        'IFERROR(INDEX(tblCategories[Joint Split A],'
        'MATCH(@{Category},tblCategories[Category],0)),DefaultSplitA))))'
    ),
    "Share A": 'IF(@{Date}="","",ROUND(@{Amount}*N(@{Split A %}),2))',
    "Share B": 'IF(@{Date}="","",ROUND(@{Amount}-N(@{Share A}),2))',
    # One column decides whose money the whole workbook reports on, so every
    # other sheet can simply sum tblTxn[View Amount].
    "View Amount": (
        'IF(@{Date}="","",IF(ReportView=PersonA,N(@{Share A}),'
        'IF(ReportView=PersonB,N(@{Share B}),@{Amount})))'
    ),
}

# The column every report adds up: the amount, or one person's share of it.
VIEW = "tblTxn[View Amount]"

TXN_FIRST_ROW = 6          # header row of tblTxn
LEDGER_CAPACITY = 20000    # rows covered by validation and formatting

PLACEHOLDER = re.compile(r"@\{([^}]+)\}")


def col_of(header: str) -> str:
    """Column letter of a transactions column (the table starts in column B)."""
    return get_column_letter(2 + TXN_HEADERS.index(header))


def col_number(header: str) -> int:
    return 2 + TXN_HEADERS.index(header)


def formula_a1(header: str, row: int) -> str:
    """A calculated column's formula as it appears on one ledger row."""
    def swap(match):
        return f"${col_of(match.group(1))}{row}"
    return "=" + PLACEHOLDER.sub(swap, TXN_FORMULAS[header])


def formula_r1c1(header: str) -> str:
    """The same formula in R1C1, which is row-position independent.

    The macros assign this to a whole block of new rows at once, so it has to
    mean "this row" wherever it lands.
    """
    def swap(match):
        return f"RC{col_number(match.group(1))}"
    return PLACEHOLDER.sub(swap, TXN_FORMULAS[header])


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------


def put(ws: Worksheet, ref: str, value, font: Font = None, fmt: str = None,
        align: str = None, fill: PatternFill = None, wrap: bool = False):
    cell = ws[ref]
    cell.value = value
    if font:
        cell.font = font
    if fmt:
        cell.number_format = fmt
    if align or wrap:
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill:
        cell.fill = fill
    return cell


def section(ws: Worksheet, row: int, first_col: str, last_col: str, title: str):
    ws.merge_cells(f"{first_col}{row}:{last_col}{row}")
    cell = ws[f"{first_col}{row}"]
    cell.value = "  " + title
    cell.font = SECTION_FONT
    cell.fill = HEAD_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20


def widths(ws: Worksheet, mapping: dict):
    for letter, width in mapping.items():
        ws.column_dimensions[letter].width = width


def add_table(ws: Worksheet, name: str, ref: str, style: str = "TableStyleMedium3"):
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=style, showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)
    return table


def validate(ws: Worksheet, ranges: Iterable[str], source: str,
             allow_blank: bool = True, prompt: str = None):
    dv = DataValidation(type="list", formula1=source, allow_blank=allow_blank,
                        showErrorMessage=False)
    if prompt:
        dv.prompt = prompt
        dv.showInputMessage = True
    ws.add_data_validation(dv)
    for ref in ranges:
        dv.add(ref)
    return dv


def name(wb: Workbook, key: str, refers_to: str):
    wb.defined_names[key] = DefinedName(key, attr_text=refers_to)


def quoted(sheet_name: str) -> str:
    return f"'{sheet_name}'" if " " in sheet_name else sheet_name


# ---------------------------------------------------------------------------
# Builder
# ---------------------------------------------------------------------------


def report_month(today: date) -> str:
    """The most recent complete month - what the Dashboard opens on."""
    return (today.replace(day=1) - timedelta(days=1)).strftime("%Y-%m")


def build(today: Optional[date] = None) -> Workbook:
    today = today or date.today()
    records = sample.build(today)

    wb = Workbook()
    wb.remove(wb.active)
    wb.code_name = "ThisWorkbook"

    # Dated from the build rather than the clock, so building the same source
    # twice produces the same bytes.
    wb.properties.creator = APP_NAME
    wb.properties.lastModifiedBy = APP_NAME
    wb.properties.title = APP_NAME
    wb.properties.description = (
        "Personal finance tracker for Canadian households. Import your bank and "
        "credit card CSV exports and read your income and expenses by month.")
    wb.properties.created = datetime.combine(today, time())
    wb.properties.modified = wb.properties.created

    build_dashboard(wb, report_month(today))
    build_transactions(wb, records)
    build_accounts(wb)
    build_categories(wb)
    build_rules(wb)
    build_budget(wb)
    build_reports(wb)
    build_household(wb)
    build_tax(wb)
    build_registered(wb)
    build_formats(wb)
    build_log(wb)
    build_settings(wb, today)
    build_help(wb)
    build_engine(wb)

    add_names(wb)
    add_validation(wb)
    add_charts(wb)

    for sheet_name, code_name in CODE_NAMES.items():
        ws = wb[sheet_name]
        ws.sheet_properties.codeName = code_name
        if sheet_name in TAB_COLOURS:
            ws.sheet_properties.tabColor = TAB_COLOURS[sheet_name]

    wb.active = wb.index(wb[SH_DASHBOARD])
    return wb


# --- Dashboard --------------------------------------------------------------

# Row positions the macros need to address as well: ApplyMode hides the couple
# block in single mode, so the named range and the block have to be built from
# the same arithmetic.
DASH_GROUPS_HEAD = 16
DASH_MERCHANTS_TOP = DASH_GROUPS_HEAD + len(data.SPENDING_GROUPS) + 2
DASH_COUPLE_TOP = DASH_MERCHANTS_TOP + 13
DASH_COUPLE_LAST = DASH_COUPLE_TOP + 10

DASH_KPIS_LEFT = [
    ("Money in", 'SUMIFS(tblTxn[View Amount],tblTxn[Month],ReportMonth,tblTxn[Type],"Income")', MONEY),
    ("Money out", '-SUMIFS(tblTxn[View Amount],tblTxn[Month],ReportMonth,tblTxn[Type],"Expense")', MONEY),
    ("Saved & invested", '-SUMIFS(tblTxn[View Amount],tblTxn[Month],ReportMonth,tblTxn[Type],"Saving")', MONEY),
    ("Net cash flow", "C9-C10-C11", MONEY),
    ("Savings rate", 'IFERROR((C9-C10)/C9,"")', PERCENT),
]

DASH_KPIS_RIGHT = [
    ("Essential spending", '-SUMIFS(tblTxn[View Amount],tblTxn[Month],ReportMonth,tblTxn[Type],"Expense",tblTxn[Essential],"Yes")', MONEY),
    ("Discretionary spending", "C10-F9", MONEY),
    ("Average spend per day", 'IFERROR(C10/DAY(EOMONTH(DATE(VALUE(LEFT(ReportMonth,4)),VALUE(RIGHT(ReportMonth,2)),1),0)),"")', MONEY),
    ("Transactions this month", "COUNTIFS(tblTxn[Month],ReportMonth)", "#,##0"),
    # Blank until the user actually sets budgets, rather than claiming they are
    # thousands of dollars over a budget of zero.
    ("Left in monthly budget",
     'IF(SUM(tblCategories[Monthly Budget])=0,"",'
     'SUM(tblCategories[Monthly Budget])-C10)', MONEY),
]


def build_dashboard(wb: Workbook, opening_month: str):
    ws = wb.create_sheet(SH_DASHBOARD)
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 2, "B": 26, "C": 15, "D": 2, "E": 26, "F": 15, "G": 2,
                "H": 22, "I": 13, "J": 2})
    for letter in "KLMNOPQRSTUVW":
        ws.column_dimensions[letter].width = 9

    ws.row_dimensions[1].height = 28
    put(ws, "B1", "Canadian Finance Tracker", TITLE_FONT)
    put(ws, "B2", "Import your bank and credit card exports, then read your month "
                  "here. Everything stays in this file - nothing is uploaded.", SUB_FONT)
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 22

    put(ws, "B6", "Report month", BOLD)
    put(ws, "C6", opening_month, LABEL_FONT, align="center",
        fill=PatternFill("solid", fgColor="FFF3CD")).border = BOX
    put(ws, "E6", "View", BOLD)
    put(ws, "F6", "Household", LABEL_FONT, align="center",
        fill=PatternFill("solid", fgColor="FFF3CD")).border = BOX
    put(ws, "H6", "Needs a category", BOLD)
    put(ws, "I6", '=COUNTIF(tblTxn[Category],"Uncategorized")', KPI_FONT,
        fmt="#,##0", align="center")

    section(ws, 8, "B", "I", "This month at a glance")
    for offset, (label, formula, fmt) in enumerate(DASH_KPIS_LEFT):
        row = 9 + offset
        put(ws, f"B{row}", label, LABEL_FONT)
        put(ws, f"C{row}", "=" + formula, KPI_FONT, fmt=fmt, align="right")
    for offset, (label, formula, fmt) in enumerate(DASH_KPIS_RIGHT):
        row = 9 + offset
        put(ws, f"E{row}", label, LABEL_FONT)
        put(ws, f"F{row}", "=" + formula, KPI_FONT, fmt=fmt, align="right")

    put(ws, "H9", "Uncategorized rows distort every number on this page. "
                  "Use the button above to clear them.", NOTE_FONT, wrap=True)
    ws.merge_cells("H9:I13")

    section(ws, DASH_GROUPS_HEAD - 1, "B", "I", "Where the money went")
    put(ws, f"B{DASH_GROUPS_HEAD}", "Group", BOLD)
    put(ws, f"C{DASH_GROUPS_HEAD}", "This month", BOLD, align="right")
    put(ws, f"E{DASH_GROUPS_HEAD}", "12-month average", BOLD, align="right")
    for offset, group in enumerate(data.SPENDING_GROUPS):
        row = DASH_GROUPS_HEAD + 1 + offset
        put(ws, f"B{row}", group, LABEL_FONT)
        put(ws, f"C{row}",
            f'=-SUMIFS(tblTxn[View Amount],tblTxn[Month],ReportMonth,'
            f'tblTxn[Group],$B{row})', fmt=MONEY, align="right")
        put(ws, f"E{row}",
            f'=-SUMIFS(tblTxn[View Amount],tblTxn[Group],$B{row},'
            f'tblTxn[Month],">="&TEXT(EDATE(DATE(VALUE(LEFT(ReportMonth,4)),'
            f'VALUE(RIGHT(ReportMonth,2)),1),-11),"yyyy-mm"),tblTxn[Month],'
            f'"<="&ReportMonth)/12', fmt=MONEY, align="right")
    last_group_row = DASH_GROUPS_HEAD + len(data.SPENDING_GROUPS)
    ws.conditional_formatting.add(
        f"C{DASH_GROUPS_HEAD + 1}:C{last_group_row}",
        DataBarRule(start_type="num", start_value=0, end_type="max", color=RED),
    )

    merchants_top = DASH_MERCHANTS_TOP
    section(ws, merchants_top, "B", "I", "Biggest merchants this month")
    put(ws, f"B{merchants_top + 1}", "Merchant", BOLD)
    put(ws, f"C{merchants_top + 1}", "Spent", BOLD, align="right")
    for row in range(merchants_top + 2, merchants_top + 12):
        ws[f"C{row}"].number_format = MONEY
    put(ws, f"E{merchants_top + 2}",
        "This list is written by the Refresh button (it needs macros).",
        NOTE_FONT, wrap=True)
    ws.merge_cells(f"E{merchants_top + 2}:I{merchants_top + 4}")

    couple_top = DASH_COUPLE_TOP
    section(ws, couple_top, "B", "I", "Couple view (selected month)")
    put(ws, f"B{couple_top + 1}", "", BOLD)
    put(ws, f"C{couple_top + 1}", "=PersonA", BOLD, align="right")
    put(ws, f"D{couple_top + 1}", "", BOLD)
    put(ws, f"E{couple_top + 1}", "=PersonB", BOLD, align="right")
    rows = [
        ("Paid from own accounts",
         '-SUMIFS(tblTxn[Amount],tblTxn[Month],ReportMonth,tblTxn[Type],"Expense",tblTxn[Paid By],{person})'),
        ("Fair share of that spending", "{share}"),
        ("Balance (paid minus share)", "{balance}"),
    ]
    shared_share_a = (
        '-(SUMIFS(tblTxn[Share A],tblTxn[Month],ReportMonth,tblTxn[Type],"Expense",tblTxn[Paid By],PersonA)'
        '+SUMIFS(tblTxn[Share A],tblTxn[Month],ReportMonth,tblTxn[Type],"Expense",tblTxn[Paid By],PersonB))'
    )
    shared_share_b = (
        '-(SUMIFS(tblTxn[Share B],tblTxn[Month],ReportMonth,tblTxn[Type],"Expense",tblTxn[Paid By],PersonA)'
        '+SUMIFS(tblTxn[Share B],tblTxn[Month],ReportMonth,tblTxn[Type],"Expense",tblTxn[Paid By],PersonB))'
    )
    row = couple_top + 2
    put(ws, f"B{row}", rows[0][0], LABEL_FONT)
    put(ws, f"C{row}", "=" + rows[0][1].format(person="PersonA"), fmt=MONEY, align="right")
    put(ws, f"E{row}", "=" + rows[0][1].format(person="PersonB"), fmt=MONEY, align="right")
    put(ws, f"B{row + 1}", rows[1][0], LABEL_FONT)
    put(ws, f"C{row + 1}", "=" + shared_share_a, fmt=MONEY, align="right")
    put(ws, f"E{row + 1}", "=" + shared_share_b, fmt=MONEY, align="right")
    put(ws, f"B{row + 2}", rows[2][0], BOLD)
    put(ws, f"C{row + 2}", f"=C{row}-C{row + 1}", BOLD, fmt=MONEY, align="right")
    put(ws, f"E{row + 2}", f"=E{row}-E{row + 1}", BOLD, fmt=MONEY, align="right")
    put(ws, f"B{row + 4}", "Settlement", BOLD)
    put(ws, f"C{row + 4}",
        f'=IF(ROUND(C{row + 2},2)>0,PersonB&" owes "&PersonA&" "&TEXT(ROUND(C{row + 2},2),"$#,##0.00"),'
        f'IF(ROUND(C{row + 2},2)<0,PersonA&" owes "&PersonB&" "&TEXT(-ROUND(C{row + 2},2),"$#,##0.00"),'
        f'"Even for this month"))', Font(size=12, bold=True, color=ACCENT))
    ws.merge_cells(f"C{row + 4}:I{row + 4}")
    put(ws, f"B{row + 6}",
        "Only spending paid from personal accounts is settled - joint accounts are "
        "assumed to be funded per the household split already. Change who an expense "
        "is for with the Owner column on the Transactions sheet.", NOTE_FONT, wrap=True)
    ws.merge_cells(f"B{row + 6}:I{row + 8}")

    ws.freeze_panes = "B7"

    # Anchors used by the macros.
    ws["B3"].value = None
    ws["B33"].value = None


# --- Transactions -----------------------------------------------------------


def build_transactions(wb: Workbook, records: Sequence[sample.Txn]):
    ws = wb.create_sheet(SH_TXN)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    for header, width in TXN_WIDTHS.items():
        ws.column_dimensions[col_of(header)].width = width

    put(ws, "B1", "Transactions", TITLE_FONT)
    put(ws, "B2", "One row per transaction. Money out is negative, money in is "
                  "positive. Grey columns are calculated - type in the white ones.",
        SUB_FONT)
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 22

    header_row = TXN_FIRST_ROW
    for index, header in enumerate(TXN_HEADERS):
        cell = ws.cell(row=header_row, column=2 + index, value=header)
        cell.font = Font(bold=True, color=WHITE)

    for offset, record in enumerate(records):
        row = header_row + 1 + offset
        write_ledger_row(ws, row, offset + 1, record)
    if not records:
        write_blank_ledger_row(ws, header_row + 1)

    last_row = header_row + max(len(records), 1)
    first_letter = col_of(TXN_HEADERS[0])
    last_letter = col_of(TXN_HEADERS[-1])
    add_table(ws, "tblTxn", f"{first_letter}{header_row}:{last_letter}{last_row}")

    # Formatting for rows the macros will add later.
    for header, fmt in (("Date", DATE_FMT), ("Amount", MONEY), ("Share A", MONEY),
                        ("Share B", MONEY), ("View Amount", MONEY),
                        ("Split A %", "0%")):
        letter = col_of(header)
        for row in range(header_row + 1, header_row + LEDGER_CAPACITY):
            ws[f"{letter}{row}"].number_format = fmt

    for header in TXN_HIDDEN:
        ws.column_dimensions[col_of(header)].hidden = True

    amount_letter = col_of("Amount")
    ws.conditional_formatting.add(
        f"{amount_letter}{header_row + 1}:{amount_letter}{header_row + LEDGER_CAPACITY}",
        CellIsRule(operator="greaterThan", formula=["0"],
                   font=Font(color="1B7F4B", bold=True)),
    )
    # Tints the whole row, so it has to test the Category cell rather than the
    # cell being formatted: a cellIs rule over these columns would only ever
    # colour Category itself.
    category_letter = col_of("Category")
    ws.conditional_formatting.add(
        f"{first_letter}{header_row + 1}:{last_letter}{header_row + LEDGER_CAPACITY}",
        FormulaRule(formula=[f'${category_letter}{header_row + 1}="Uncategorized"'],
                    fill=PatternFill("solid", fgColor="FDE8E8"),
                    stopIfTrue=False),
    )
    ws.conditional_formatting.add(
        f"{category_letter}{header_row + 1}:{category_letter}{header_row + LEDGER_CAPACITY}",
        CellIsRule(operator="equal", formula=['"Uncategorized"'],
                   font=Font(color=RED_DARK, bold=True)),
    )

    ws.freeze_panes = f"D{header_row + 1}"


def write_ledger_row(ws: Worksheet, row: int, sequence: int, record: sample.Txn):
    values = {
        "Txn ID": f"T{sequence:06d}",
        "Date": record.when,
        "Account": record.account,
        "Owner": record.owner,
        "Description": record.description,
        "Merchant": record.merchant,
        "Amount": record.amount,
        "Category": record.category,
        "Source File": "sample-data",
        "Batch": "SAMPLE",
        "Match Key": sample.match_key(record.account, record.when, record.amount,
                                      record.description),
        "Tagged By": "Rule",
    }
    for header in TXN_HEADERS:
        cell = ws.cell(row=row, column=col_number(header))
        if header in TXN_FORMULAS:
            cell.value = formula_a1(header, row)
        elif header in values:
            cell.value = values[header]
    ws.cell(row=row, column=col_number("Date")).number_format = DATE_FMT
    ws.cell(row=row, column=col_number("Amount")).number_format = MONEY


def write_blank_ledger_row(ws: Worksheet, row: int):
    for header in TXN_HEADERS:
        if header in TXN_FORMULAS:
            ws.cell(row=row, column=col_number(header),
                    value=formula_a1(header, row))


# --- Reference sheets -------------------------------------------------------


def build_accounts(wb: Workbook):
    ws = wb.create_sheet(SH_ACCOUNTS)
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 2, "B": 26, "C": 20, "D": 14, "E": 14, "F": 28, "G": 24,
                "H": 18, "I": 30})

    put(ws, "B1", "Accounts", TITLE_FONT)
    put(ws, "B2", "One row per bank or card account. \"File Name Contains\" lets the "
                  "importer recognise a download without asking.", SUB_FONT)

    headers = ["Account", "Institution", "Type", "Owner", "Bank Format",
               "File Name Contains", "Include in Household", "Notes"]
    for index, header in enumerate(headers):
        ws.cell(row=4, column=2 + index, value=header).font = Font(bold=True, color=WHITE)

    rows = [
        (sample.ACCOUNT_ALEX, "RBC Royal Bank", "Chequing", sample.PERSON_A,
         "RBC Chequing/Savings/Card", "rbc-chequing-alex", "Yes", "Sample account"),
        (sample.ACCOUNT_SAM, "Tangerine", "Chequing", sample.PERSON_B,
         "Tangerine", "tangerine-chequing-sam", "Yes", "Sample account"),
        (sample.ACCOUNT_JOINT, "BMO", "Chequing", "Joint", "BMO",
         "bmo-joint-chequing", "Yes", "Sample account"),
        (sample.ACCOUNT_CARD, "American Express", "Credit Card", "Joint",
         "Amex Canada", "amex-cobalt-joint", "Yes", "Sample account"),
        ("", "", "", "", "", "", "", ""),
        ("", "", "", "", "", "", "", ""),
        ("", "", "", "", "", "", "", ""),
        ("", "", "", "", "", "", "", ""),
    ]
    for offset, row_values in enumerate(rows):
        for index, value in enumerate(row_values):
            ws.cell(row=5 + offset, column=2 + index, value=value or None)

    add_table(ws, "tblAccounts", f"B4:I{4 + len(rows)}")
    ws.freeze_panes = "B5"


def build_categories(wb: Workbook):
    ws = wb.create_sheet(SH_CATEGORIES)
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 2, "B": 28, "C": 16, "D": 11, "E": 11, "F": 26, "G": 15,
                "H": 14, "I": 13, "J": 34})

    put(ws, "B1", "Categories", TITLE_FONT)
    put(ws, "B2", "Add your own rows freely. Type drives the reports: Income and "
                  "Expense count in cash flow, Saving is money you kept, Transfer is "
                  "ignored. \"Default Owner\" decides whose expense it is when the "
                  "rules categorise a row, and \"Joint Split A\" overrides the "
                  "household split for one category.", SUB_FONT)

    for index, header in enumerate(data.CATEGORY_COLUMNS):
        ws.cell(row=4, column=2 + index, value=header).font = Font(bold=True, color=WHITE)

    for offset, entry in enumerate(data.CATEGORIES):
        category, group, kind, essential, tag, budget, notes = entry
        row = 5 + offset
        ws.cell(row=row, column=2, value=category)
        ws.cell(row=row, column=3, value=group)
        ws.cell(row=row, column=4, value=kind)
        ws.cell(row=row, column=5, value=essential)
        ws.cell(row=row, column=6, value=tag or None)
        budget_cell = ws.cell(row=row, column=7, value=budget or None)
        budget_cell.number_format = MONEY
        ws.cell(row=row, column=8, value=data.default_owner(category) or None)
        split_cell = ws.cell(row=row, column=9, value="=DefaultSplitA")
        split_cell.number_format = "0%"
        ws.cell(row=row, column=10, value=notes or None)

    last = 4 + len(data.CATEGORIES)
    add_table(ws, "tblCategories", f"B4:J{last}")
    ws.freeze_panes = "B5"


def build_rules(wb: Workbook):
    ws = wb.create_sheet(SH_RULES)
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 2, "B": 9, "C": 9, "D": 13, "E": 13, "F": 32, "G": 12,
                "H": 12, "I": 12, "J": 26, "K": 12, "L": 8, "M": 26})

    put(ws, "B1", "Rules", TITLE_FONT)
    put(ws, "B2", "Rules run in Priority order, lowest first, and the first match "
                  "wins. The Teach a rule button on the Transactions sheet adds new "
                  "ones at priority 10.", SUB_FONT)

    for index, header in enumerate(data.RULE_COLUMNS):
        ws.cell(row=4, column=2 + index, value=header).font = Font(bold=True, color=WHITE)

    rules = data.seed_rules()
    for offset, rule in enumerate(rules):
        for index, value in enumerate(rule):
            cell = ws.cell(row=5 + offset, column=2 + index,
                           value=value if value != "" else None)
            if index in (5, 6):
                cell.number_format = MONEY

    last = 4 + len(rules)
    add_table(ws, "tblRules", f"B4:M{last}")
    ws.freeze_panes = "B5"


def build_formats(wb: Workbook):
    ws = wb.create_sheet(SH_FORMATS)
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 2, "B": 32, "C": 22, "D": 10, "E": 11, "F": 9, "G": 14,
                "H": 17, "I": 14, "J": 10, "K": 10, "L": 11, "M": 34, "N": 60})

    put(ws, "B1", "Bank formats", TITLE_FONT)
    put(ws, "B2", "How each bank's CSV is laid out. Column numbers are 1-based. "
                  "If an import looks wrong, fix the numbers here - no code changes "
                  "needed. Rows that cannot be read as a date (headers, notices) are "
                  "skipped automatically.", SUB_FONT)

    for index, header in enumerate(data.FORMAT_COLUMNS):
        ws.cell(row=4, column=2 + index, value=header).font = Font(bold=True, color=WHITE)

    for offset, row_values in enumerate(data.BANK_FORMATS):
        for index, value in enumerate(row_values):
            ws.cell(row=5 + offset, column=2 + index, value=value)

    last = 4 + len(data.BANK_FORMATS)
    add_table(ws, "tblFormats", f"B4:N{last}")
    ws.freeze_panes = "B5"


def build_log(wb: Workbook):
    ws = wb.create_sheet(SH_LOG)
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 2, "B": 19, "C": 22, "D": 30, "E": 28, "F": 24, "G": 11,
                "H": 11, "I": 12, "J": 12})

    put(ws, "B1", "Import log", TITLE_FONT)
    put(ws, "B2", "Written by the importer: one row per file.", SUB_FONT)

    headers = ["When", "Batch", "File", "Format", "Account", "Rows read",
               "Imported", "Duplicates", "Unreadable"]
    for index, header in enumerate(headers):
        ws.cell(row=4, column=2 + index, value=header).font = Font(bold=True, color=WHITE)
    ws.cell(row=5, column=2, value=None)

    add_table(ws, "tblLog", "B4:J5")
    ws.freeze_panes = "B5"


# --- Reports ----------------------------------------------------------------

REPORT_MONTHS = 12
REPORT_FIRST_COL = 3          # column C
REPORT_CATEGORY_ROWS = 120

CASHFLOW_ROWS = [
    ("Money in", '=SUMIFS(tblTxn[View Amount],tblTxn[Month],{month},tblTxn[Type],"Income")', MONEY),
    ("Money out", '=-SUMIFS(tblTxn[View Amount],tblTxn[Month],{month},tblTxn[Type],"Expense")', MONEY),
    ("Saved & invested", '=-SUMIFS(tblTxn[View Amount],tblTxn[Month],{month},tblTxn[Type],"Saving")', MONEY),
    ("Net cash flow", "={col}{income}-{col}{spend}-{col}{saved}", MONEY),
    ("Savings rate", '=IFERROR(({col}{income}-{col}{spend})/{col}{income},"")', PERCENT),
    ("Essential spending", '=-SUMIFS(tblTxn[View Amount],tblTxn[Month],{month},tblTxn[Type],"Expense",tblTxn[Essential],"Yes")', MONEY),
    ("Discretionary spending", "={col}{spend}-{col}{essential}", MONEY),
    ("Transactions", "=COUNTIFS(tblTxn[Month],{month})", "#,##0"),
]


def build_reports(wb: Workbook):
    ws = wb.create_sheet(SH_REPORTS)
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 2, "B": 30})
    for index in range(REPORT_MONTHS + 1):
        ws.column_dimensions[get_column_letter(REPORT_FIRST_COL + index)].width = 12
    total_col = get_column_letter(REPORT_FIRST_COL + REPORT_MONTHS)
    ws.column_dimensions[total_col].width = 14
    sign_col = get_column_letter(REPORT_FIRST_COL + REPORT_MONTHS + 1)
    ws.column_dimensions[sign_col].hidden = True

    put(ws, "B1", "Reports", TITLE_FONT)
    put(ws, "B2", "Twelve months ending with the report month chosen on the "
                  "Dashboard, in the same view (household or one person).", SUB_FONT)

    header_row = 4
    put(ws, f"B{header_row}", "Month", Font(bold=True, color=WHITE))
    for index in range(REPORT_MONTHS):
        letter = get_column_letter(REPORT_FIRST_COL + index)
        offset = index - (REPORT_MONTHS - 1)
        put(ws, f"{letter}{header_row}",
            f'=TEXT(EDATE(DATE(VALUE(LEFT(ReportMonth,4)),VALUE(RIGHT(ReportMonth,2)),1),'
            f'{offset}),"yyyy-mm")', Font(bold=True, color=WHITE), align="center")
    put(ws, f"{total_col}{header_row}", "Total", Font(bold=True, color=WHITE),
        align="center")
    for column in range(2, REPORT_FIRST_COL + REPORT_MONTHS + 1):
        ws.cell(row=header_row, column=column).fill = HEAD_FILL

    first_row = header_row + 1
    income_row = first_row
    spend_row = first_row + 1
    saved_row = first_row + 2
    essential_row = first_row + 5

    for offset, (label, template, fmt) in enumerate(CASHFLOW_ROWS):
        row = first_row + offset
        put(ws, f"B{row}", label, BOLD if "Net" in label else LABEL_FONT)
        for index in range(REPORT_MONTHS):
            letter = get_column_letter(REPORT_FIRST_COL + index)
            formula = template.format(
                month=f"{letter}${header_row}", col=letter, income=income_row,
                spend=spend_row, saved=saved_row, essential=essential_row,
            )
            put(ws, f"{letter}{row}", formula, fmt=fmt, align="right")
        first_letter = get_column_letter(REPORT_FIRST_COL)
        last_letter = get_column_letter(REPORT_FIRST_COL + REPORT_MONTHS - 1)
        if fmt == PERCENT:
            put(ws, f"{total_col}{row}",
                f"=IFERROR(({total_col}{income_row}-{total_col}{spend_row})/"
                f"{total_col}{income_row},\"\")", fmt=fmt, align="right")
        else:
            put(ws, f"{total_col}{row}",
                f"=SUM({first_letter}{row}:{last_letter}{row})", BOLD, fmt=fmt,
                align="right")

    group_head = first_row + len(CASHFLOW_ROWS) + 1
    section(ws, group_head, "B", total_col, "Spending by group")
    group_first = group_head + 1
    for offset, group in enumerate(data.SPENDING_GROUPS):
        row = group_first + offset
        put(ws, f"B{row}", group, LABEL_FONT)
        for index in range(REPORT_MONTHS):
            letter = get_column_letter(REPORT_FIRST_COL + index)
            put(ws, f"{letter}{row}",
                f'=-SUMIFS(tblTxn[View Amount],tblTxn[Month],{letter}${header_row},'
                f'tblTxn[Group],$B{row})', fmt=MONEY, align="right")
        put(ws, f"{total_col}{row}",
            f"=SUM({get_column_letter(REPORT_FIRST_COL)}{row}:"
            f"{get_column_letter(REPORT_FIRST_COL + REPORT_MONTHS - 1)}{row})",
            BOLD, fmt=MONEY, align="right")
    group_last = group_first + len(data.SPENDING_GROUPS) - 1

    category_head = group_last + 2
    section(ws, category_head, "B", total_col, "By category (income positive, "
                                               "spending positive)")
    category_first = category_head + 1
    for offset in range(REPORT_CATEGORY_ROWS):
        row = category_first + offset
        put(ws, f"B{row}",
            f'=IFERROR(INDEX(tblCategories[Category],{offset + 1}),"")', LABEL_FONT)
        put(ws, f"{sign_col}{row}",
            f'=IF($B{row}="",0,IFERROR(IF(INDEX(tblCategories[Type],'
            f'MATCH($B{row},tblCategories[Category],0))="Income",1,-1),-1))')
        for index in range(REPORT_MONTHS):
            letter = get_column_letter(REPORT_FIRST_COL + index)
            put(ws, f"{letter}{row}",
                f'=IF($B{row}="","",SUMIFS(tblTxn[View Amount],tblTxn[Month],'
                f'{letter}${header_row},tblTxn[Category],$B{row})*${sign_col}{row})',
                fmt=MONEY, align="right")
        put(ws, f"{total_col}{row}",
            f'=IF($B{row}="","",SUM({get_column_letter(REPORT_FIRST_COL)}{row}:'
            f'{get_column_letter(REPORT_FIRST_COL + REPORT_MONTHS - 1)}{row}))',
            BOLD, fmt=MONEY, align="right")

    ws.freeze_panes = f"{get_column_letter(REPORT_FIRST_COL)}{header_row + 1}"


# --- Budget -----------------------------------------------------------------

BUDGET_ROWS = 120


def build_budget(wb: Workbook):
    ws = wb.create_sheet(SH_BUDGET)
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 2, "B": 28, "C": 15, "D": 10, "E": 14, "F": 14, "G": 14,
                "H": 11, "I": 16, "J": 26})
    ws.column_dimensions["K"].hidden = True

    put(ws, "B1", "Budget", TITLE_FONT)
    put(ws, "B2", "Set a monthly budget per category on the Categories sheet; this "
                  "page compares it with what actually happened in the report month "
                  "chosen on the Dashboard. Income and spending are both shown as "
                  "positive numbers, and a negative Difference always means the month "
                  "went the wrong way.", SUB_FONT)
    put(ws, "B3", "Month shown", BOLD)
    put(ws, "C3", "=ReportMonth", Font(bold=True, color=ACCENT), align="center")

    headers = ["Category", "Group", "Type", "Budget", "Actual", "Difference",
               "% used", "12-month average", "Notes"]
    for index, header in enumerate(headers):
        cell = ws.cell(row=5, column=2 + index, value=header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = HEAD_FILL

    def lookup(row: int, column: str) -> str:
        return (f'IFERROR(INDEX(tblCategories[{column}],'
                f'MATCH($B{row},tblCategories[Category],0)),"")')

    for offset in range(BUDGET_ROWS):
        row = 6 + offset
        put(ws, f"B{row}", f'=IFERROR(INDEX(tblCategories[Category],{offset + 1}),"")')
        put(ws, f"C{row}", f'=IF($B{row}="","",{lookup(row, "Group")})')
        put(ws, f"D{row}", f'=IF($B{row}="","",{lookup(row, "Type")})')
        # Money in counts up, money out counts down: one sign puts both on the
        # same footing so "more than budgeted" always reads the same way.
        put(ws, f"K{row}", f'=IF($D{row}="Income",1,-1)')
        put(ws, f"E{row}",
            f'=IF($B{row}="","",IFERROR(INDEX(tblCategories[Monthly Budget],'
            f'MATCH($B{row},tblCategories[Category],0)),0))', fmt=MONEY)
        put(ws, f"F{row}",
            f'=IF($B{row}="","",SUMIFS(tblTxn[View Amount],tblTxn[Month],'
            f'ReportMonth,tblTxn[Category],$B{row})*$K{row})', fmt=MONEY)
        put(ws, f"G{row}",
            f'=IF($B{row}="","",IF($D{row}="Income",F{row}-E{row},E{row}-F{row}))',
            fmt=MONEY)
        put(ws, f"H{row}", f'=IF(N(E{row})=0,"",F{row}/E{row})', fmt=PERCENT)
        put(ws, f"I{row}",
            f'=IF($B{row}="","",SUMIFS(tblTxn[View Amount],tblTxn[Category],'
            f'$B{row},tblTxn[Month],">="&TEXT(EDATE(DATE(VALUE(LEFT(ReportMonth,4)),'
            f'VALUE(RIGHT(ReportMonth,2)),1),-11),"yyyy-mm"),tblTxn[Month],'
            f'"<="&ReportMonth)*$K{row}/12)', fmt=MONEY)

    last = 5 + BUDGET_ROWS
    ws.conditional_formatting.add(
        f"G6:G{last}",
        CellIsRule(operator="lessThan", formula=["0"],
                   font=Font(color=RED_DARK, bold=True)),
    )
    ws.conditional_formatting.add(
        f"H6:H{last}",
        ColorScaleRule(start_type="num", start_value=0, start_color="C6EFCE",
                       mid_type="num", mid_value=1, mid_color="FFEB9C",
                       end_type="num", end_value=1.5, end_color="FFC7CE"),
    )
    ws.freeze_panes = "B6"


# --- Household --------------------------------------------------------------


def build_household(wb: Workbook):
    ws = wb.create_sheet(SH_HOUSEHOLD)
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 2, "B": 36, "C": 16, "D": 16, "E": 16, "F": 4, "G": 44})

    put(ws, "B1", "Household", TITLE_FONT)
    put(ws, "B2", "Who paid what, whose it was, and who owes whom. Uses the report "
                  "month chosen on the Dashboard.", SUB_FONT)
    put(ws, "B3", "Month shown", BOLD)
    put(ws, "C3", "=ReportMonth", Font(bold=True, color=ACCENT), align="center")

    section(ws, 5, "B", "E", "Settling up (personal accounts only)")
    put(ws, "B6", "", BOLD)
    put(ws, "C6", "=PersonA", BOLD, align="right")
    put(ws, "D6", "=PersonB", BOLD, align="right")
    put(ws, "E6", "Together", BOLD, align="right")

    paid = ('-SUMIFS(tblTxn[Amount],tblTxn[Month],ReportMonth,tblTxn[Type],'
            '"Expense",tblTxn[Paid By],{person})')
    share = ('-(SUMIFS(tblTxn[{col}],tblTxn[Month],ReportMonth,tblTxn[Type],"Expense",'
             'tblTxn[Paid By],PersonA)+SUMIFS(tblTxn[{col}],tblTxn[Month],ReportMonth,'
             'tblTxn[Type],"Expense",tblTxn[Paid By],PersonB))')

    put(ws, "B7", "Paid from own accounts", LABEL_FONT)
    put(ws, "C7", "=" + paid.format(person="PersonA"), fmt=MONEY, align="right")
    put(ws, "D7", "=" + paid.format(person="PersonB"), fmt=MONEY, align="right")
    put(ws, "E7", "=C7+D7", BOLD, fmt=MONEY, align="right")

    put(ws, "B8", "Fair share of that spending", LABEL_FONT)
    put(ws, "C8", "=" + share.format(col="Share A"), fmt=MONEY, align="right")
    put(ws, "D8", "=" + share.format(col="Share B"), fmt=MONEY, align="right")
    put(ws, "E8", "=C8+D8", BOLD, fmt=MONEY, align="right")

    put(ws, "B9", "Balance (paid minus share)", BOLD)
    put(ws, "C9", "=C7-C8", BOLD, fmt=MONEY, align="right")
    put(ws, "D9", "=D7-D8", BOLD, fmt=MONEY, align="right")

    put(ws, "B11", "Settlement", BOLD)
    put(ws, "C11",
        '=IF(ROUND(C9,2)>0,PersonB&" owes "&PersonA&" "&TEXT(ROUND(C9,2),"$#,##0.00"),'
        'IF(ROUND(C9,2)<0,PersonA&" owes "&PersonB&" "&TEXT(-ROUND(C9,2),"$#,##0.00"),'
        '"Even for this month"))', Font(size=12, bold=True, color=ACCENT))
    ws.merge_cells("C11:E11")

    section(ws, 13, "B", "E", "Everything, including joint accounts")
    put(ws, "B14", "Household spending", LABEL_FONT)
    put(ws, "C14", '=-SUMIFS(tblTxn[Amount],tblTxn[Month],ReportMonth,tblTxn[Type],"Expense")',
        fmt=MONEY, align="right")
    put(ws, "B15", "Your share of it", LABEL_FONT)
    put(ws, "C15", '=-SUMIFS(tblTxn[Share A],tblTxn[Month],ReportMonth,tblTxn[Type],"Expense")',
        fmt=MONEY, align="right")
    put(ws, "D15", '=-SUMIFS(tblTxn[Share B],tblTxn[Month],ReportMonth,tblTxn[Type],"Expense")',
        fmt=MONEY, align="right")
    put(ws, "B16", "Income received", LABEL_FONT)
    put(ws, "C16", '=SUMIFS(tblTxn[Amount],tblTxn[Month],ReportMonth,tblTxn[Type],"Income",tblTxn[Owner],PersonA)',
        fmt=MONEY, align="right")
    put(ws, "D16", '=SUMIFS(tblTxn[Amount],tblTxn[Month],ReportMonth,tblTxn[Type],"Income",tblTxn[Owner],PersonB)',
        fmt=MONEY, align="right")
    put(ws, "E16", "=C16+D16", BOLD, fmt=MONEY, align="right")
    put(ws, "B17", "Share of household income", LABEL_FONT)
    put(ws, "C17", '=IFERROR(C16/$E$16,"")', fmt=PERCENT, align="right")
    put(ws, "D17", '=IFERROR(D16/$E$16,"")', fmt=PERCENT, align="right")
    put(ws, "B18", "Split currently used for joint costs", LABEL_FONT)
    put(ws, "C18", "=DefaultSplitA", fmt=PERCENT, align="right")
    put(ws, "D18", "=1-DefaultSplitA", fmt=PERCENT, align="right")
    put(ws, "B19", "Income-proportional split would be", LABEL_FONT)
    put(ws, "C19", '=IFERROR(C16/$E$16,"")', fmt=PERCENT, align="right")
    put(ws, "D19", '=IFERROR(D16/$E$16,"")', fmt=PERCENT, align="right")

    section(ws, 21, "B", "E", "Year to date")
    year_criteria = 'LEFT(ReportMonth,4)&"-*"'
    put(ws, "B22", "Household spending", LABEL_FONT)
    put(ws, "C22", f'=-SUMIFS(tblTxn[Amount],tblTxn[Month],{year_criteria},tblTxn[Type],"Expense")',
        fmt=MONEY, align="right")
    put(ws, "B23", "Your share of it", LABEL_FONT)
    put(ws, "C23", f'=-SUMIFS(tblTxn[Share A],tblTxn[Month],{year_criteria},tblTxn[Type],"Expense")',
        fmt=MONEY, align="right")
    put(ws, "D23", f'=-SUMIFS(tblTxn[Share B],tblTxn[Month],{year_criteria},tblTxn[Type],"Expense")',
        fmt=MONEY, align="right")
    put(ws, "B24", "Balance so far", BOLD)
    put(ws, "C24",
        f'=-SUMIFS(tblTxn[Amount],tblTxn[Month],{year_criteria},tblTxn[Type],"Expense",'
        f'tblTxn[Paid By],PersonA)+SUMIFS(tblTxn[Share A],tblTxn[Month],{year_criteria},'
        f'tblTxn[Type],"Expense",tblTxn[Paid By],PersonA)+SUMIFS(tblTxn[Share A],'
        f'tblTxn[Month],{year_criteria},tblTxn[Type],"Expense",tblTxn[Paid By],PersonB)',
        BOLD, fmt=MONEY, align="right")

    put(ws, "G5", "How the split works", BOLD)
    put(ws, "G6",
        "Every transaction has an Owner: one of the two of you, or Joint.\n\n"
        "- Owner = a person: that person carries 100% of it.\n"
        "- Owner = Joint: it is divided using the household split on the Settings "
        "sheet, unless the category overrides it in \"Joint Split A\".\n\n"
        "\"Paid By\" comes from the account the money left, so a shared dinner put on "
        "one person's card shows up as that person paying more than their share.\n\n"
        "Settling up only looks at personal accounts. Spending from a joint account "
        "is assumed to be funded by both of you already.\n\n"
        "Switch between household and per-person figures with the View selector on "
        "the Dashboard.", LABEL_FONT, wrap=True)
    ws.merge_cells("G6:G24")


# --- Tax summary ------------------------------------------------------------


def build_tax(wb: Workbook):
    ws = wb.create_sheet(SH_TAX)
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 2, "B": 30, "C": 16, "D": 16, "E": 16, "F": 58})

    put(ws, "B1", "Tax summary", TITLE_FONT)
    put(ws, "B2", "Totals for anything tagged on the Categories sheet. This is a "
                  "record of what you spent, not tax advice - keep your receipts and "
                  "check the rules that apply to you.", SUB_FONT)
    put(ws, "B3", "Tax year", BOLD)
    put(ws, "C3", "=VALUE(LEFT(ReportMonth,4))", Font(bold=True, color=ACCENT),
        fmt="0", align="center")

    headers = ["Tax tag", "Total", "=PersonA", "=PersonB", "Where it usually goes"]
    for index, header in enumerate(headers):
        cell = ws.cell(row=5, column=2 + index, value=header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = HEAD_FILL

    for offset, (tag, note) in enumerate(data.TAX_TAGS):
        row = 6 + offset
        put(ws, f"B{row}", tag, LABEL_FONT)
        put(ws, f"C{row}",
            f'=ABS(SUMIFS(tblTxn[Amount],tblTxn[Tax Tag],$B{row},tblTxn[Month],'
            f'TaxYear&"-*"))', fmt=MONEY, align="right")
        put(ws, f"D{row}",
            f'=ABS(SUMIFS(tblTxn[Share A],tblTxn[Tax Tag],$B{row},tblTxn[Month],'
            f'TaxYear&"-*"))', fmt=MONEY, align="right")
        put(ws, f"E{row}",
            f'=ABS(SUMIFS(tblTxn[Share B],tblTxn[Tax Tag],$B{row},tblTxn[Month],'
            f'TaxYear&"-*"))', fmt=MONEY, align="right")
        put(ws, f"F{row}", note, NOTE_FONT, wrap=True)

    last = 5 + len(data.TAX_TAGS)
    put(ws, f"B{last + 2}",
        "Medical expenses can be claimed for any 12-month period ending in the tax "
        "year, and are usually best claimed by the lower-income spouse. Donations can "
        "be pooled on one return. Both rules are worth checking before you file.",
        NOTE_FONT, wrap=True)
    ws.merge_cells(f"B{last + 2}:F{last + 4}")
    ws.freeze_panes = "B6"


# --- Registered plans -------------------------------------------------------

PLANS = [
    ("RRSP", "RRSP Contribution", 33810),
    ("TFSA", "TFSA Contribution", 7000),
    ("FHSA", "FHSA Contribution", data.FHSA_ANNUAL_LIMIT),
    ("RESP", "RESP Contribution", data.RESP_ANNUAL_GRANT_TARGET),
]


def build_registered(wb: Workbook):
    ws = wb.create_sheet(SH_REGISTERED)
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 2, "B": 18, "C": 16, "D": 16, "E": 15, "F": 4, "G": 16,
                "H": 16, "I": 15, "J": 4, "K": 14, "L": 14, "M": 14})

    put(ws, "B1", "Registered plans", TITLE_FONT)
    put(ws, "B2", "Contributions counted from the categories RRSP/TFSA/FHSA/RESP "
                  "Contribution, for the tax year of the report month. Enter your own "
                  "room from your CRA notice of assessment - only the CRA knows your "
                  "real numbers.", SUB_FONT)

    put(ws, "C4", "=PersonA", BOLD, align="center")
    ws.merge_cells("C4:E4")
    put(ws, "G4", "=PersonB", BOLD, align="center")
    ws.merge_cells("G4:I4")

    headers = ["Plan", "Contributed", "Your room", "Left", "", "Contributed",
               "Your room", "Left"]
    positions = ["B", "C", "D", "E", "F", "G", "H", "I"]
    for letter, header in zip(positions, headers):
        if not header:
            continue
        cell = ws[f"{letter}5"]
        cell.value = header
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = HEAD_FILL

    for offset, (plan, category, default_room) in enumerate(PLANS):
        row = 6 + offset
        put(ws, f"B{row}", plan, BOLD)
        put(ws, f"C{row}",
            f'=ABS(SUMIFS(tblTxn[Amount],tblTxn[Category],"{category}",'
            f'tblTxn[Owner],PersonA,tblTxn[Month],TaxYear&"-*"))', fmt=MONEY,
            align="right")
        put(ws, f"D{row}", default_room, fmt=MONEY, align="right")
        put(ws, f"E{row}", f"=D{row}-C{row}", BOLD, fmt=MONEY, align="right")
        put(ws, f"G{row}",
            f'=ABS(SUMIFS(tblTxn[Amount],tblTxn[Category],"{category}",'
            f'tblTxn[Owner],PersonB,tblTxn[Month],TaxYear&"-*"))', fmt=MONEY,
            align="right")
        put(ws, f"H{row}", default_room, fmt=MONEY, align="right")
        put(ws, f"I{row}", f"=H{row}-G{row}", BOLD, fmt=MONEY, align="right")

    plan_rows = range(6, 6 + len(PLANS))
    put(ws, "B10", "Not counted: owner is Joint", BOLD)
    put(ws, "C10",
        "=" + "+".join(
            f'ABS(SUMIFS(tblTxn[Amount],tblTxn[Category],"{category}",'
            f'tblTxn[Owner],"Joint",tblTxn[Month],TaxYear&"-*"))'
            for _plan, category, _room in PLANS
        ), BOLD, fmt=MONEY, align="right")
    put(ws, "B11", "A registered plan always belongs to one person, so contributions "
                   "left with the Joint owner are not counted above. Set the Owner "
                   "column on the Transactions sheet to whoever's plan it is - the "
                   "\"Set owner\" button does it for a whole selection.",
        NOTE_FONT, wrap=True)
    ws.merge_cells("B11:I12")

    section(ws, 14, "B", "E", "TFSA dollar limit by year")
    put(ws, "B15", "Year", BOLD)
    put(ws, "C15", "Limit", BOLD, align="right")
    for offset, (year, limit) in enumerate(data.TFSA_LIMITS):
        row = 16 + offset
        put(ws, f"B{row}", year, fmt="0")
        put(ws, f"C{row}", limit, fmt=MONEY0, align="right")

    section(ws, 14, "G", "I", "RRSP dollar limit by year")
    put(ws, "G15", "Year", BOLD)
    put(ws, "H15", "Limit", BOLD, align="right")
    for offset, (year, limit) in enumerate(data.RRSP_LIMITS):
        row = 16 + offset
        put(ws, f"G{row}", year, fmt="0")
        put(ws, f"H{row}", limit, fmt=MONEY0, align="right")

    put(ws, "G25", "Your RRSP room is the lesser of 18% of last year's earned income "
                   "and the dollar limit, plus carry-forward, minus any pension "
                   "adjustment.", NOTE_FONT, wrap=True)
    ws.merge_cells("G25:I27")

    put(ws, "K5", "Other reference numbers", BOLD)
    reference = [
        ("FHSA annual limit", data.FHSA_ANNUAL_LIMIT),
        ("FHSA lifetime limit", data.FHSA_LIFETIME_LIMIT),
        ("RESP for full CESG", data.RESP_ANNUAL_GRANT_TARGET),
        ("RESP lifetime limit", data.RESP_LIFETIME_LIMIT),
    ]
    for offset, (label, value) in enumerate(reference):
        row = 6 + offset
        put(ws, f"K{row}", label, LABEL_FONT)
        put(ws, f"L{row}", value, fmt=MONEY0, align="right")

    put(ws, "K11", "CPP pensionable earnings", BOLD)
    put(ws, "K12", "Year", BOLD)
    put(ws, "L12", "YMPE", BOLD, align="right")
    put(ws, "M12", "YAMPE", BOLD, align="right")
    for offset, (year, ympe, yampe) in enumerate(data.PENSIONABLE_EARNINGS):
        row = 13 + offset
        put(ws, f"K{row}", year, fmt="0")
        put(ws, f"L{row}", ympe, fmt=MONEY0, align="right")
        put(ws, f"M{row}", yampe, fmt=MONEY0, align="right")

    put(ws, "K17", "Figures published by the CRA and checked in September 2026. "
                   "Confirm against canada.ca or your CRA account before you rely on "
                   "them.", NOTE_FONT, wrap=True)
    ws.merge_cells("K17:M20")


# --- Settings ---------------------------------------------------------------

SETTINGS_ROWS = [
    ("Household mode", "Couple", "Single for one person, Couple for two."),
    ("Person A", sample.PERSON_A, "Shown everywhere as the first person."),
    ("Person B", sample.PERSON_B, "Only used in Couple mode."),
    ("Person A share of joint costs", 0.5,
     "Used for transactions whose Owner is Joint. A category can override it."),
    ("Province or territory", "ON", "Used for the notes on the Tax summary sheet."),
    ("Setup completed", "No", "Set to Yes by the setup wizard."),
    ("Transfer match window (days)", 4,
     "How far apart the two sides of a transfer can be."),
    ("Skip duplicates on import", "Yes",
     "Re-importing an overlapping statement adds only the new rows."),
    ("Currency", "CAD", "Everything is treated as Canadian dollars."),
    ("Version", "1.0.0", "Workbook version."),
]


def build_settings(wb: Workbook, today: date):
    ws = wb.create_sheet(SH_SETTINGS)
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 2, "B": 34, "C": 18, "D": 2, "E": 62})

    put(ws, "B1", "Settings", TITLE_FONT)
    put(ws, "B2", "Change these here or run the setup wizard from the Dashboard.",
        SUB_FONT)

    section(ws, 4, "B", "E", "Household")
    for offset, (label, value, note) in enumerate(SETTINGS_ROWS):
        row = 5 + offset
        put(ws, f"B{row}", label, LABEL_FONT)
        cell = put(ws, f"C{row}", value, Font(bold=True, color=INK), align="center",
                   fill=PatternFill("solid", fgColor="FFF3CD"))
        cell.border = BOX
        if isinstance(value, float):
            cell.number_format = "0%"
        put(ws, f"E{row}", note, NOTE_FONT, wrap=True)

    row = 5 + len(SETTINGS_ROWS) + 1
    section(ws, row, "B", "E", "Privacy")
    put(ws, f"B{row + 1}",
        "This workbook never connects to your bank and never sends anything "
        "anywhere. It only reads the CSV files you choose. Keep the file somewhere "
        "safe - it contains your complete spending history.", LABEL_FONT, wrap=True)
    ws.merge_cells(f"B{row + 1}:E{row + 3}")


# --- Engine (hidden) --------------------------------------------------------


def build_engine(wb: Workbook):
    ws = wb.create_sheet(SH_ENGINE)
    ws.sheet_state = "hidden"
    widths(ws, {"A": 22, "B": 90, "D": 12, "F": 14, "H": 14, "J": 12, "L": 8,
                "N": 16, "P": 14, "R": 14, "T": 14, "V": 12, "X": 14, "Z": 14,
                "AB": 10, "AD": 12, "AF": 26, "AH": 16})

    put(ws, "A1", "Engine: lists and formula templates. Please do not delete this "
                  "sheet - the macros and the drop-downs depend on it.", NOTE_FONT)

    put(ws, "A3", "Column", BOLD)
    put(ws, "B3", "Formula (R1C1)", BOLD)
    for offset, header in enumerate(TXN_FORMULAS):
        row = 4 + offset
        ws[f"A{row}"] = header
        # Stored as text, without the leading "=", so Excel leaves it alone.
        ws[f"B{row}"] = formula_r1c1(header)
    add_table(ws, "tblTemplates", f"A3:B{3 + len(TXN_FORMULAS)}",
              style="TableStyleLight1")

    for offset in range(36):
        put(ws, f"D{3 + offset}", f'=TEXT(EDATE(TODAY(),-{offset}),"yyyy-mm")')

    put(ws, "F3", "Household")
    put(ws, "F4", "=PersonA")
    put(ws, "F5", "=PersonB")

    put(ws, "H3", "=PersonA")
    put(ws, "H4", "=PersonB")
    put(ws, "H5", "Joint")

    for offset, value in enumerate(data.CATEGORY_TYPES):
        put(ws, f"J{3 + offset}", value)
    put(ws, "L3", "Yes")
    put(ws, "L4", "No")
    for offset, value in enumerate(data.ACCOUNT_TYPES):
        put(ws, f"N{3 + offset}", value)
    for offset, value in enumerate(data.RULE_FIELDS):
        put(ws, f"P{3 + offset}", value)
    for offset, value in enumerate(data.RULE_TESTS):
        put(ws, f"R{3 + offset}", value)
    for offset, value in enumerate(data.RULE_FLOWS):
        put(ws, f"T{3 + offset}", value)
    for offset, value in enumerate(data.DELIMITERS):
        put(ws, f"V{3 + offset}", value)
    for offset, value in enumerate(data.AMOUNT_MODES):
        put(ws, f"X{3 + offset}", value)
    for offset, value in enumerate(data.DATE_FORMATS):
        put(ws, f"Z{3 + offset}", value)
    for offset, (code, _label) in enumerate(data.PROVINCES):
        put(ws, f"AB{3 + offset}", code)
    for offset, (tag, _note) in enumerate(data.TAX_TAGS):
        put(ws, f"AF{3 + offset}", tag)
    for offset, group in enumerate(data.GROUPS):
        put(ws, f"AH{3 + offset}", group)

    put(ws, "AD2", "Rows in the ledger", NOTE_FONT)
    put(ws, "AD3", "=COUNTA(tblTxn[Txn ID])")


# --- Help -------------------------------------------------------------------

HELP_SECTIONS = [
    ("Getting started", [
        "1. Enable macros when Excel asks - the import, categorising and buttons "
        "need them. If you downloaded this file, right-click it, choose Properties "
        "and tick Unblock first.",
        "2. Press \"Setup wizard\" on the Dashboard: it asks who the workbook is for, "
        "your names, how joint costs are split and your province, then offers to "
        "delete the sample data.",
        "3. List your accounts on the Accounts sheet - one row per bank or card "
        "account. Fill in \"File Name Contains\" with a snippet of the file name your "
        "bank produces so imports find the right account by themselves.",
        "4. Download a CSV from each bank and press \"Import statements\".",
    ]),
    ("Importing", [
        "The importer reads CSV files. Excel and PDF statements are not supported - "
        "every Canadian bank offers a CSV download from the transaction list.",
        "It recognises the layout from the file's own header line, shows you the "
        "first few rows it parsed, and asks you to confirm before anything is written.",
        "Money leaving an account is stored as a negative number and money arriving "
        "as a positive one, whichever way your bank writes it.",
        "Re-importing a statement that overlaps one you already loaded adds only the "
        "new rows. Two identical purchases on the same day are still kept as two.",
        "If a bank's columns do not match, fix the column numbers on the Bank Formats "
        "sheet - it takes a few seconds and no code changes.",
    ]),
    ("Categories and rules", [
        "Rules turn merchant names into categories. They run in Priority order and "
        "the first match wins, so specific rules should have a lower number.",
        "Anything the rules cannot place lands in Uncategorized and is highlighted in "
        "red. Press \"Needs a category\" to filter to just those rows.",
        "Select a row and press \"Teach a rule\" to create a rule from it. From then "
        "on every matching transaction is categorised automatically.",
        "A category you type in by hand is marked Manual and is never overwritten by "
        "the rules.",
        "Transfers between your own accounts and credit card payments are detected by "
        "matching opposite amounts within a few days, and are excluded from income and "
        "expenses so they do not double count.",
    ]),
    ("Couples", [
        "Set Household mode to Couple in the setup wizard or on the Settings sheet.",
        "Every transaction has an Owner: one of you, or Joint. Joint transactions are "
        "divided using the household split, which a category can override in \"Joint "
        "Split A\" on the Categories sheet.",
        "\"Paid By\" is worked out from the account the money left, so it is possible "
        "to pay for something that is not (all) yours.",
        "The Household sheet compares what each of you paid from your own accounts "
        "with your fair share, and says who owes whom.",
        "The View selector on the Dashboard switches every number between the whole "
        "household and one person's share.",
    ]),
    ("Canadian bits", [
        "Categories cover the things a Canadian household actually pays: hydro, "
        "Presto/Compass/OPUS, LCBO/SAQ, property tax, daycare, RRSP/TFSA/FHSA/RESP, "
        "OSAP, and CRA payments and refunds.",
        "Deposits like the Canada Child Benefit, the GST/HST credit and the Canada "
        "Carbon Rebate are recognised as income, not as a random transfer.",
        "The Tax summary sheet totals everything tagged as medical, donations, child "
        "care, tuition, professional dues and so on, for the tax year of the report "
        "month, split per person.",
        "The Registered plans sheet tracks RRSP/TFSA/FHSA/RESP contributions per "
        "person against the room you enter, and lists the published annual limits.",
        "None of this is tax advice: confirm your own contribution room and what you "
        "can claim with the CRA or your accountant.",
    ]),
    ("Keyboard shortcuts", [
        "Ctrl+Shift+I - import statements",
        "Ctrl+Shift+R - apply rules to uncategorised rows",
        "Ctrl+Shift+T - teach a rule from the selected row",
        "Ctrl+Shift+U - show rows that still need a category",
    ]),
    ("If something looks wrong", [
        "Numbers stale? Press Refresh on the Dashboard, or Ctrl+Alt+F9.",
        "A calculated column overwritten by accident? Press \"Rebuild formulas\" on "
        "the Transactions sheet.",
        "Buttons missing? They are drawn when the file opens with macros enabled. "
        "Close and reopen, or run modUI.EnsureButtons from the macro list.",
        "An import mapped the wrong columns? Adjust that row on the Bank Formats "
        "sheet and import again - duplicates will be skipped.",
        "Want to start over? \"Start fresh\" on the Transactions sheet empties the "
        "ledger but keeps your accounts, categories, rules and settings.",
    ]),
]


def build_help(wb: Workbook):
    ws = wb.create_sheet(SH_HELP)
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 2, "B": 4, "C": 110})

    put(ws, "B1", "How this workbook works", TITLE_FONT)
    put(ws, "B2", "Canadian Finance Tracker - everything runs locally in this file.",
        SUB_FONT)

    row = 4
    for title, lines in HELP_SECTIONS:
        section(ws, row, "B", "C", title)
        row += 1
        for line in lines:
            cell = put(ws, f"C{row}", line, LABEL_FONT, wrap=True)
            ws.row_dimensions[row].height = 15 * (1 + len(line) // 105)
            row += 1
        row += 1

    put(ws, f"C{row}",
        "Not financial or tax advice. Figures published by the CRA were checked in "
        "September 2026; confirm anything that matters against canada.ca or your CRA "
        "account.", NOTE_FONT, wrap=True)


# --- Names, validation, charts ---------------------------------------------


def add_names(wb: Workbook):
    settings = quoted(SH_SETTINGS)
    engine = quoted(SH_ENGINE)
    dash = quoted(SH_DASHBOARD)

    keys = ["HouseholdMode", "PersonA", "PersonB", "DefaultSplitA", "Province",
            "Configured", "TransferWindowDays", "SkipDuplicates", "Currency",
            "AppVersion"]
    for offset, key in enumerate(keys):
        name(wb, key, f"{settings}!$C${5 + offset}")

    name(wb, "ReportMonth", f"{dash}!$C$6")
    name(wb, "ReportView", f"{dash}!$F$6")
    name(wb, "ButtonAnchor", f"{dash}!$B$3")
    name(wb, "TxnButtonAnchor", f"{quoted(SH_TXN)}!$B$3")

    name(wb, "TopMerchants",
         f"{dash}!$B${DASH_MERCHANTS_TOP + 2}:$C${DASH_MERCHANTS_TOP + 11}")
    name(wb, "CoupleBlock",
         f"{dash}!$B${DASH_COUPLE_TOP}:$I${DASH_COUPLE_LAST}")

    name(wb, "MonthList", f"{engine}!$D$3:$D$38")
    name(wb, "ViewList", f"{engine}!$F$3:$F$5")
    name(wb, "HouseholdOnly", f"{engine}!$F$3")
    name(wb, "OwnerList", f"{engine}!$H$3:$H$5")
    name(wb, "TypeList", f"{engine}!$J$3:$J${2 + len(data.CATEGORY_TYPES)}")
    name(wb, "YesNoList", f"{engine}!$L$3:$L$4")
    name(wb, "AccountTypeList", f"{engine}!$N$3:$N${2 + len(data.ACCOUNT_TYPES)}")
    name(wb, "RuleFieldList", f"{engine}!$P$3:$P${2 + len(data.RULE_FIELDS)}")
    name(wb, "RuleTestList", f"{engine}!$R$3:$R${2 + len(data.RULE_TESTS)}")
    name(wb, "RuleFlowList", f"{engine}!$T$3:$T${2 + len(data.RULE_FLOWS)}")
    name(wb, "DelimiterList", f"{engine}!$V$3:$V${2 + len(data.DELIMITERS)}")
    name(wb, "AmountModeList", f"{engine}!$X$3:$X${2 + len(data.AMOUNT_MODES)}")
    name(wb, "DateFormatList", f"{engine}!$Z$3:$Z${2 + len(data.DATE_FORMATS)}")
    name(wb, "ProvinceList", f"{engine}!$AB$3:$AB${2 + len(data.PROVINCES)}")
    name(wb, "LedgerRows", f"{engine}!$AD$3")
    name(wb, "TaxTagList", f"{engine}!$AF$3:$AF${2 + len(data.TAX_TAGS)}")
    name(wb, "GroupList", f"{engine}!$AH$3:$AH${2 + len(data.GROUPS)}")
    name(wb, "TaxYear", f"{quoted(SH_TAX)}!$C$3")

    name(wb, "CategoryList", "=tblCategories[Category]")
    name(wb, "AccountList", "=tblAccounts[Account]")
    name(wb, "FormatList", "=tblFormats[Profile]")


def add_validation(wb: Workbook):
    ws = wb[SH_TXN]
    first = TXN_FIRST_ROW + 1
    last = TXN_FIRST_ROW + LEDGER_CAPACITY
    validate(ws, [f"{col_of('Category')}{first}:{col_of('Category')}{last}"],
             "=CategoryList")
    validate(ws, [f"{col_of('Account')}{first}:{col_of('Account')}{last}"],
             "=AccountList")
    validate(ws, [f"{col_of('Owner')}{first}:{col_of('Owner')}{last}"], "=OwnerList")
    validate(ws, [f"{col_of('Reimbursable')}{first}:{col_of('Reimbursable')}{last}"],
             "=YesNoList")

    ws = wb[SH_DASHBOARD]
    validate(ws, ["C6"], "=MonthList", prompt="Pick a month to report on")
    validate(ws, ["F6"], "=ViewList",
             prompt="Household, or one person's share")

    ws = wb[SH_ACCOUNTS]
    validate(ws, ["D5:D200"], "=AccountTypeList")
    validate(ws, ["E5:E200"], "=OwnerList")
    validate(ws, ["F5:F200"], "=FormatList")
    validate(ws, ["H5:H200"], "=YesNoList")

    ws = wb[SH_CATEGORIES]
    last_category = 4 + len(data.CATEGORIES) + 100
    validate(ws, [f"C5:C{last_category}"], "=GroupList")
    validate(ws, [f"D5:D{last_category}"], "=TypeList")
    validate(ws, [f"E5:E{last_category}"], "=YesNoList")
    validate(ws, [f"F5:F{last_category}"], "=TaxTagList")
    validate(ws, [f"H5:H{last_category}"], "=OwnerList")

    ws = wb[SH_RULES]
    last_rule = 4 + len(data.seed_rules()) + 200
    validate(ws, [f"C5:C{last_rule}"], "=YesNoList")
    validate(ws, [f"D5:D{last_rule}"], "=RuleFieldList")
    validate(ws, [f"E5:E{last_rule}"], "=RuleTestList")
    validate(ws, [f"I5:I{last_rule}"], "=RuleFlowList")
    validate(ws, [f"J5:J{last_rule}"], "=CategoryList")
    validate(ws, [f"K5:K{last_rule}"], "=OwnerList")

    ws = wb[SH_FORMATS]
    last_format = 4 + len(data.BANK_FORMATS) + 50
    validate(ws, [f"E5:E{last_format}"], "=DelimiterList")
    validate(ws, [f"G5:G{last_format}"], "=DateFormatList")
    validate(ws, [f"I5:I{last_format}"], "=AmountModeList")

    ws = wb[SH_SETTINGS]
    validate(ws, ["C5"], '"Single,Couple"')
    validate(ws, ["C9"], "=ProvinceList")
    validate(ws, ["C10"], '"Yes,No"')
    validate(ws, ["C12"], '"Yes,No"')


def add_charts(wb: Workbook):
    dashboard = wb[SH_DASHBOARD]
    reports = wb[SH_REPORTS]

    header_row = 4
    income_row = 5
    spend_row = 6
    net_row = 8
    group_first = 5 + len(CASHFLOW_ROWS) + 2
    group_last = group_first + len(data.SPENDING_GROUPS) - 1

    bar = BarChart()
    bar.type = "col"
    bar.title = "Money in and money out"
    bar.height = 7.5
    bar.width = 17
    bar.y_axis.numFmt = "#,##0"
    bar.gapWidth = 60
    values = Reference(reports, min_col=2, max_col=2 + REPORT_MONTHS,
                       min_row=income_row, max_row=spend_row)
    categories = Reference(reports, min_col=REPORT_FIRST_COL,
                           max_col=REPORT_FIRST_COL + REPORT_MONTHS - 1,
                           min_row=header_row, max_row=header_row)
    bar.add_data(values, titles_from_data=True, from_rows=True)
    bar.set_categories(categories)
    dashboard.add_chart(bar, "K3")

    line = LineChart()
    line.title = "Net cash flow"
    line.height = 7.5
    line.width = 17
    line.y_axis.numFmt = "#,##0"
    net = Reference(reports, min_col=2, max_col=2 + REPORT_MONTHS,
                    min_row=net_row, max_row=net_row)
    line.add_data(net, titles_from_data=True, from_rows=True)
    line.set_categories(categories)
    dashboard.add_chart(line, "K19")

    doughnut = DoughnutChart()
    doughnut.title = "Where the money went"
    doughnut.height = 8.5
    doughnut.width = 17
    labels = Reference(dashboard, min_col=2, min_row=17, max_row=16 + len(data.SPENDING_GROUPS))
    amounts = Reference(dashboard, min_col=3, min_row=16,
                        max_row=16 + len(data.SPENDING_GROUPS))
    doughnut.add_data(amounts, titles_from_data=True)
    doughnut.set_categories(labels)
    dashboard.add_chart(doughnut, "K35")
